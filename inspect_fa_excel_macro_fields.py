from pathlib import Path
import json, re
from openpyxl import load_workbook
ROOT=Path('.').resolve()
files=sorted(ROOT.glob('MWG_valuation_model*.xlsx'))+sorted(ROOT.glob('MWG_SOTP_Scenarios.xlsx'))
terms=['GDP','CPI','lãi suất','lai suat','interest','deposit','WACC','risk-free','risk free','tỷ giá','ty gia','USD','VND','retail','bán lẻ','ban le','SSSG','inflation','PMI','credit','tín dụng','tin dung','bond','yield','market risk premium','ERP','discount rate','cost of debt']
out=[]
for f in files:
    try:
        wb=load_workbook(f, data_only=False, read_only=True)
    except Exception as e:
        out.append({'file':str(f),'error':str(e)}); continue
    hits=[]
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=min(ws.max_row or 0,200), max_col=min(ws.max_column or 0,50)):
            vals=[]
            for c in row:
                v=c.value
                if isinstance(v,(str,int,float)):
                    vals.append(str(v))
            line=' | '.join(vals)
            low=line.lower()
            if any(t.lower() in low for t in terms):
                hits.append({'sheet':ws.title,'row':row[0].row if row else None,'text':line[:500]})
    out.append({'file':str(f),'sheets':wb.sheetnames,'hits':hits[:80], 'hit_count':len(hits)})
Path('macro_data_inventory/fa_excel_macro_hits.json').write_text(json.dumps(out,ensure_ascii=False,indent=2),encoding='utf-8')
for item in out:
    print('\nFILE',item['file'],'hits',item.get('hit_count'))
    for h in item.get('hits',[])[:20]:
        print(' ',h['sheet'],h['row'],h['text'][:220])
