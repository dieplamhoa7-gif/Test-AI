from openpyxl import load_workbook
from pathlib import Path

src=Path('MWG_valuation_model_v4_fixed.xlsx')
out=Path('MWG_valuation_model_v5_data_quality.xlsx')
wb=load_workbook(src)

# Make data quality very explicit and stop pretending segment 2025 exactness
# 1) Add banner notes on key sheets
for shname in ['Historical_5Y','Segment_Historical','TGDD_Forecast','DMX_Forecast','BHX_Forecast','Dashboard']:
    ws=wb[shname]
    ws.insert_rows(1,2)
    ws['A1']='WARNING'
    ws['B1']='Các số segment hiện tại là structured estimate để model chạy logic. Chưa được tie-out 100% với annual report / segment note chính thức của MWG.'
    ws['A2']='STATUS'
    ws['B2']='Consolidated structure usable; segment exact 2025 still pending official tie-out.'

# 2) Update dashboard status text
ws=wb['Dashboard']
ws['A16']='Data warning'
ws['B16']='Target price hiện chỉ có ý nghĩa kịch bản/phương pháp. Không dùng như fair value cuối cho đến khi tie-out xong số lịch sử và segment 2025.'

# 3) Strengthen Data_Quality sheet
ws=wb['Data_Quality']
rows=[
('Official 2025 consolidated PDF','Downloaded from CafeF list and stored locally','High','Need line-item extraction / manual key-in from audited PDF'),
('Official 2025 segment revenue/profit','Not yet extracted cleanly from official disclosures','Low','Need annual report / investor presentation / segment note'),
('Current model usage','Best used for framework, scenario analysis, and sensitivity','High','Do not present as final audited valuation yet'),
]
start=ws.max_row+1
for i,r in enumerate(rows,start):
    for j,v in enumerate(r,1): ws.cell(i,j).value=v

# 4) Add a new sheet requesting exact data inputs explicitly
if 'Required_Exact_Data' in wb.sheetnames:
    del wb['Required_Exact_Data']
ws=wb.create_sheet('Required_Exact_Data')
ws.append(['Need exact item','Why needed','Suggested source','Priority'])
items=[
('2021-2025 TGDD revenue','Needed to tie history and revenue/store','Annual report / investor presentation','High'),
('2021-2025 ĐMX revenue','Core earnings driver','Annual report / investor presentation','High'),
('2021-2025 BHX revenue','Core grocery forecast base','Annual report / investor presentation','High'),
('2021-2025 BHX EBIT margin / loss','Critical for turnaround path','Annual report / management commentary','High'),
('2021-2025 store count TGDD/ĐMX/BHX','Needed for revenue/store logic','Monthly KPI / annual report','High'),
('ĐMX market share by year','Needed for share-driven forecast','Company commentary / industry report','Medium'),
('BHX market share by year','Needed for grocery penetration case','Industry report / management commentary','Medium'),
('Current net debt and shares','Needed for final value/share','Latest financial statements','High'),
]
for r in items: ws.append(r)
for col in ['A','B','C','D']: ws.column_dimensions[col].width=36

wb.save(out)
print(out.resolve())
