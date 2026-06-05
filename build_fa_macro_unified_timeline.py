from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date
import json, csv, re, math

BASE=Path('FA')
OUT=BASE/'data'/'unified_macro'
OUT.mkdir(parents=True, exist_ok=True)

DATE_RE=re.compile(r'^(Q[1-4]/\d{4}|\d{1,2}[-/]\d{4}|\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{4})$')

def norm(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    s=str(v).strip().replace('\n',' ')
    return re.sub(r'\s+',' ',s)

def is_dateish(v):
    if isinstance(v,(datetime,date)): return True
    s=norm(v)
    if not s: return False
    if DATE_RE.match(s): return True
    # Excel Kỳ style like 3-2026
    if re.match(r'^\d{1,2}-\d{4}$',s): return True
    return False

def clean_indicator(s):
    s=norm(s)
    s=re.sub(r'\s*Đơn vị:.*$','',s,flags=re.I)
    return s.strip()

def to_float(v):
    if v is None or v=='': return None
    if isinstance(v,(int,float)) and not isinstance(v,bool):
        if isinstance(v,float) and (math.isnan(v) or math.isinf(v)): return None
        return float(v)
    s=str(v).strip().replace(',','')
    try: return float(s)
    except Exception: return None

def classify_indicator(name, file):
    low=(name+' '+file).lower()
    if any(x in low for x in ['omo','nghiệp vụ thị trường mở','thi truong mo','tín phiếu','tin phieu','bơm','bom','hút','hut']): return 'liquidity_omo'
    if any(x in low for x in ['lãi suất','lai suat','overnight','qua đêm','tái cấp vốn','tái chiết khấu','huy động']): return 'rates'
    if any(x in low for x in ['usd','eur','cny','tỷ giá','ty gia','fx']): return 'fx'
    if any(x in low for x in ['m2','tiền gửi','tin dung','tín dụng','dư nợ']): return 'money_credit'
    if any(x in low for x in ['xuất khẩu','xuat khau','nhập khẩu','nhap khau','cán cân','can can','f.o.b','current account','vãng lai']): return 'trade_bop'
    if any(x in low for x in ['gdp','cpi','iip','pmi','fdi','bán lẻ','ban le','retail']): return 'macro_growth_inflation'
    return 'other_macro'

def extract_file(path:Path):
    rows=[]
    wb=load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        data=[]
        for r in ws.iter_rows(values_only=True):
            data.append(list(r))
        if not data: continue
        # Pattern A: row table with header containing Ngày and many indicator columns.
        for hi,row in enumerate(data[:15]):
            vals=[norm(x) for x in row]
            if any(v.lower()=='ngày' or v.lower()=='ngay' for v in vals) and len([v for v in vals if v])>=4:
                date_idx=next(i for i,v in enumerate(vals) if v.lower() in ('ngày','ngay'))
                headers=vals
                for rr in data[hi+1:]:
                    d=norm(rr[date_idx] if date_idx<len(rr) else '')
                    if not d or not is_dateish(d): continue
                    for ci,h in enumerate(headers):
                        if ci<=date_idx or not h: continue
                        val=rr[ci] if ci<len(rr) else None
                        fv=to_float(val)
                        if fv is None: continue
                        ind=clean_indicator(h)
                        rows.append({'date':d,'indicator':ind,'value':fv,'source_file':path.name,'sheet':ws.title,'category':classify_indicator(ind,path.name),'orientation':'row_table'})
                break
        # Pattern B: wide table, first nonempty row has date-ish headers, first col indicators.
        for hi,row in enumerate(data[:12]):
            headers=[norm(x) for x in row]
            date_cols=[i for i,h in enumerate(headers) if is_dateish(h)]
            if len(date_cols)>=2:
                for rr in data[hi+1:]:
                    ind=clean_indicator(rr[0] if rr else '')
                    if not ind or ind.lower() in ('stt','kỳ','ngày'): continue
                    for ci in date_cols:
                        val=rr[ci] if ci<len(rr) else None
                        fv=to_float(val)
                        if fv is None: continue
                        d=headers[ci]
                        rows.append({'date':d,'indicator':ind,'value':fv,'source_file':path.name,'sheet':ws.title,'category':classify_indicator(ind,path.name),'orientation':'wide_table'})
                break
    return rows

all_rows=[]
for f in sorted(BASE.glob('*.xlsx')):
    all_rows.extend(extract_file(f))

# de-duplicate exact duplicates
seen=set(); unique=[]
for r in all_rows:
    key=(r['date'],r['indicator'],r['value'],r['source_file'],r['sheet'])
    if key in seen: continue
    seen.add(key); unique.append(r)
unique.sort(key=lambda r:(str(r['date']), r['category'], r['source_file'], r['indicator']))

fields=['date','category','indicator','value','source_file','sheet','orientation']
with (OUT/'macro_timeline_unified.csv').open('w',encoding='utf-8-sig',newline='') as f:
    w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(unique)
(OUT/'macro_timeline_unified.json').write_text(json.dumps({'createdAt':datetime.now().isoformat(),'rowCount':len(unique),'rows':unique},ensure_ascii=False,indent=2),encoding='utf-8')

# Coverage vs code/registry
registry=json.loads((BASE/'data'/'source_registry.json').read_text(encoding='utf-8'))
reg_fields=[]
for s in registry.get('sources',[]):
    for field in s.get('data_fields',[]): reg_fields.append({'source':s['id'],'field':str(field)})
fetcher_text='\n'.join(p.read_text(encoding='utf-8',errors='ignore') for p in (BASE/'macro'/'fetchers').glob('*.py'))
fetcher_low=fetcher_text.lower()
indicators=sorted(set(r['indicator'] for r in unique))
coverage=[]
for ind in indicators:
    low=ind.lower()
    tokens=[t for t in re.split(r'[^a-zA-Z0-9À-ỹ]+',low) if len(t)>=4]
    hit_registry=[]
    for rf in reg_fields:
        if low in rf['field'].lower() or rf['field'].lower() in low:
            hit_registry.append(rf['source']+': '+rf['field'])
    # conservative code hit if key words appear
    code_hit=any(t in fetcher_low for t in tokens[:4]) if tokens else False
    status='covered_or_partial' if hit_registry or code_hit else 'not_found_in_code'
    coverage.append({'indicator':ind,'category':classify_indicator(ind,''),'timeline_rows':sum(1 for r in unique if r['indicator']==ind),'registry_matches':hit_registry[:5],'code_keyword_hit':code_hit,'status':status})
coverage.sort(key=lambda x:(x['status'],x['category'],x['indicator']))
(OUT/'macro_code_coverage.json').write_text(json.dumps(coverage,ensure_ascii=False,indent=2),encoding='utf-8')

summary=[]
summary.append('# Unified Macro Timeline + Code Coverage\n')
summary.append('Scope: all Excel macro files directly under `FA/`.\n')
summary.append(f'- Unified timeline rows: **{len(unique)}**')
summary.append(f'- Unique indicators: **{len(indicators)}**')
summary.append(f'- Output CSV: `FA/data/unified_macro/macro_timeline_unified.csv`')
summary.append(f'- Output JSON: `FA/data/unified_macro/macro_timeline_unified.json`')
summary.append(f'- Code coverage JSON: `FA/data/unified_macro/macro_code_coverage.json`\n')
summary.append('## Rows by category\n')
for cat in sorted(set(r['category'] for r in unique)):
    summary.append(f'- {cat}: {sum(1 for r in unique if r["category"]==cat)} rows; {len(set(r["indicator"] for r in unique if r["category"]==cat))} indicators')
summary.append('\n## Code coverage status\n')
for st in sorted(set(c['status'] for c in coverage)):
    summary.append(f'- {st}: {sum(1 for c in coverage if c["status"]==st)} indicators')
summary.append('\n## Important not/weakly covered items to verify\n')
for c in coverage:
    if c['status']=='not_found_in_code' and c['category']!='other_macro':
        summary.append(f'- `{c["indicator"]}` ({c["category"]}) — {c["timeline_rows"]} rows')
summary.append('\n## Notes\n')
summary.append('- This is an automated first-pass parser. It preserves source_file/sheet for audit. Human validation is still required before using in production scoring.')
summary.append('- Existing code has fetchers for Pinetree, VCB FX, yfinance global, vnstock market, WorldBank, SBV rates, SBV OMO. However a field is only production-ready after a successful run populates `FA/data/history/YYYY-MM-DD.json` with non-null data and matching schema.')
(OUT/'UNIFIED_MACRO_TIMELINE_REPORT.md').write_text('\n'.join(summary),encoding='utf-8')

print(json.dumps({'rows':len(unique),'indicators':len(indicators),'out':str(OUT)},ensure_ascii=False,indent=2))
