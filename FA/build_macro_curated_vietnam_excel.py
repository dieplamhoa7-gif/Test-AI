from __future__ import annotations

import csv, json, re
from datetime import datetime
from pathlib import Path
from typing import Any
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parent
OUT=ROOT/'reports'; OUT.mkdir(exist_ok=True)

IMPORTANT_RULES = [
    # Liquidity/rates - VN core
    ('Thanh khoản & lãi suất', 'Lãi suất liên ngân hàng qua đêm', ['interbankOvernight','liên nh qua đêm','overnight','qua đêm'], ['usd','eur']),
    ('Thanh khoản & lãi suất', 'Lãi suất liên ngân hàng 1 tuần', ['interbank1W','1w','1 tuần'], ['usd']),
    ('Thanh khoản & lãi suất', 'Lãi suất liên ngân hàng 1 tháng', ['interbank1M','1m','1 tháng'], ['usd']),
    ('Thanh khoản & lãi suất', 'Lãi suất OMO / Reverse Repo', ['omoRate','lãi suất omo','reverse repo'], []),
    ('Thanh khoản & lãi suất', 'Bơm/hút ròng OMO', ['totalLiquidityNetBn','reverseRepoNetBn','bơm','hút','omo'], []),
    ('Thanh khoản & lãi suất', 'Lãi suất huy động 12 tháng', ['deposit12m','tiết kiệm 12','huy động 12'], []),
    ('Thanh khoản & lãi suất', 'Tín phiếu NHNN lưu hành/phát hành', ['tín phiếu','tbill'], []),
    ('Thanh khoản & lãi suất', 'M2 / cung tiền', ['moneySupplyM2','m2','cung tiền'], []),
    ('Thanh khoản & lãi suất', 'Tăng trưởng tín dụng', ['credit','tín dụng'], []),
    # FX/external
    ('Tỷ giá & đối ngoại', 'USD/VND', ['usdVnd','usd/vnd','vnd-usd','usd-vnd','usd vnd'], []),
    ('Tỷ giá & đối ngoại', 'DXY', ['dxy','dollar index'], []),
    ('Tỷ giá & đối ngoại', 'Dự trữ ngoại hối', ['fx reserves','foreign exchange reserves','dự trữ ngoại hối'], []),
    ('Tỷ giá & đối ngoại', 'Cán cân thương mại', ['trade balance','cán cân thương mại'], []),
    ('Tỷ giá & đối ngoại', 'Xuất khẩu', ['exports','xuất khẩu'], ['% gdp']),
    ('Tỷ giá & đối ngoại', 'Nhập khẩu', ['imports','nhập khẩu'], ['% gdp']),
    ('Tỷ giá & đối ngoại', 'Cán cân vãng lai', ['current account','cán cân vãng lai'], []),
    ('Tỷ giá & đối ngoại', 'FDI giải ngân / thực hiện', ['fdi','vốn thực hiện','disbursed'], []),
    # Growth/inflation
    ('Tăng trưởng & lạm phát', 'GDP tăng trưởng', ['gdp growth','gdp','tăng trưởng gdp'], ['% gdp','current us','nominal']),
    ('Tăng trưởng & lạm phát', 'CPI / lạm phát', ['inflation','cpi','lạm phát'], []),
    ('Tăng trưởng & lạm phát', 'CPI lương thực/thực phẩm', ['food cpi','lương thực','thực phẩm'], []),
    ('Tăng trưởng & lạm phát', 'Bán lẻ', ['retail','bán lẻ'], []),
    ('Tăng trưởng & lạm phát', 'IIP / sản xuất công nghiệp', ['industrial production','iip','sản xuất công nghiệp'], []),
    ('Tăng trưởng & lạm phát', 'PMI', ['pmi'], []),
    # Market/risk
    ('Thị trường & khẩu vị rủi ro', 'VNINDEX', ['vnindex','vn-index'], []),
    ('Thị trường & khẩu vị rủi ro', 'Thanh khoản thị trường chứng khoán', ['marketTurnoverBn','gtgd','thanh khoản','turnover'], []),
    ('Thị trường & khẩu vị rủi ro', 'Khối ngoại mua/bán ròng', ['foreignNetBuyBn','khối ngoại','nđtnn','foreign'], []),
    ('Thị trường & khẩu vị rủi ro', 'VIX', ['vix'], []),
    ('Thị trường & khẩu vị rủi ro', 'S&P 500', ['sp500','s&p'], []),
    ('Thị trường & khẩu vị rủi ro', 'US 10Y', ['us10y','us 10y','treasury','dgs10'], []),
    ('Hàng hóa', 'Giá dầu Brent', ['brent'], []),
    ('Hàng hóa', 'Giá vàng', ['gold','vàng'], []),
]

DROP_PATTERNS = [
    'eurvnd','cnyvnd','jpy','gbp','aud','cad','chf','hkd','sgd','thb','krw',
    'djia','nasdaq','ftse','dax','cac40','hnx','vn30','upcom',
    'moneySupplyM0','moneySupplyM1','m0','m1',
]


def load_csv(p:Path):
    if not p.exists(): return []
    with p.open('r',encoding='utf-8-sig',newline='') as f: return list(csv.DictReader(f))

def num(v):
    if v is None or v=='': return None
    try: return float(str(v).replace(',',''))
    except: return None

def norm(s): return re.sub(r'\s+',' ',str(s or '').strip())

def keytext(*xs): return ' '.join(norm(x).lower() for x in xs if x)

def period_rank(s):
    s=str(s or '')
    m=re.match(r'^(\d{4})-(\d{2})-(\d{2})$',s)
    if m: return (int(m.group(1)),int(m.group(2)),int(m.group(3)))
    m=re.match(r'^(\d{1,2})-(\d{4})$',s)
    if m: return (int(m.group(2)),int(m.group(1)),28)
    m=re.match(r'^Q([1-4])/(\d{4})$',s,re.I)
    if m: return (int(m.group(2)),int(m.group(1))*3,30)
    m=re.search(r'(20\d{2}|19\d{2})',s)
    if m: return (int(m.group(1)),12,31)
    return (0,0,0)

def classify(period, source_file=''):
    s=str(period or ''); sf=str(source_file or '').lower()
    if 'can can' in sf or re.match(r'^Q[1-4]/\d{4}$',s,re.I): return 'Quý'
    if any(x in sf for x in ['du_lieu_vi_mo','lai suat huy dong','lai suat thong ke','20266']) or re.match(r'^\d{1,2}-\d{4}$',s): return 'Tháng'
    if re.match(r'^\d{4}$',s): return 'Năm'
    if re.match(r'^\d{4}-\d{2}-\d{2}$',s): return 'Ngày'
    return 'Khác'

def collect_rows():
    rows=[]
    # Pinetree
    for r in load_csv(ROOT/'data/pinetree_archive/pinetree_macro_timeline.csv'):
        v=num(r.get('value'))
        if v is None: continue
        rows.append({'moc':r.get('date'),'tan_suat':'Ngày','chi_so_goc':norm(r.get('label') or r.get('indicator')),'schema':r.get('indicator'),'gia_tri':v,'nguon':'Pinetree Bản tin sáng','file_nguon':'FA/data/pinetree_archive/pinetree_macro_timeline.csv','note':'Daily/trading-day; nguồn public, tự cập nhật được qua crawler.'})
    # Unified macro/FiinProX
    for r in load_csv(ROOT/'data/unified_macro/macro_timeline_unified.csv'):
        v=num(r.get('value'))
        if v is None: continue
        rows.append({'moc':r.get('date'),'tan_suat':classify(r.get('date'),r.get('source_file')),'chi_so_goc':norm(r.get('indicator')),'schema':'','gia_tri':v,'nguon':'FiinProX Excel/manual backfill','file_nguon':r.get('source_file'),'note':'Manual/fallback/backfill; không coi là nguồn daily chính.'})
    # TradingEconomics visible history/latest
    te_hist=ROOT/'data/tradingeconomics_visible_history.csv'
    te_src=te_hist if te_hist.exists() else ROOT/'data/tradingeconomics_visible_timeline.csv'
    for r in load_csv(te_src):
        v=num(r.get('value') or r.get('last'))
        if v is None: continue
        moc=r.get('fetchedDate') or (r.get('fetchedAt') or '')[:10]
        rows.append({'moc':moc,'tan_suat':'Ngày','chi_so_goc':norm(r.get('indicator')),'schema':r.get('key'),'gia_tri':v,'nguon':'TradingEconomics visible','file_nguon':str(te_src.relative_to(ROOT)),'note':'Daily latest visible snapshot; không dùng paid/download.'})
    # WorldBank
    for p in reversed(sorted((ROOT/'data/history').glob('2026-06-05_v*.json'))):
        try:
            snap=json.loads(p.read_text(encoding='utf-8')); wb=snap.get('worldbank',{}).get('data',{})
            if wb: break
        except: wb={}
    for k,obj in (wb or {}).items():
        for yr,v in (obj.get('timeSeries') or {}).items():
            if v is None: continue
            rows.append({'moc':yr,'tan_suat':'Năm','chi_so_goc':norm(k),'schema':k,'gia_tri':v,'nguon':'WorldBank API','file_nguon':str(p.relative_to(ROOT)),'note':'Annual/lagged; chỉ dùng làm bối cảnh dài hạn.'})
    # SBV liquidity latest
    liq=ROOT/'data/sbv_liquidity/latest.json'
    if liq.exists():
        try:
            d=json.loads(liq.read_text(encoding='utf-8')); s=d.get('summary',{}); moc=s.get('date') or (d.get('fetchedAt') or '')[:10]
            for label,key in [('Bơm/hút ròng OMO', 'reverseRepoNetBn'),('Reverse Repo phát hành','reverseRepoIssueBn'),('Lãi suất OMO / Reverse Repo','omoRate'),('Bơm/hút ròng toàn hệ thống NHNN','totalLiquidityNetBn')]:
                v=s.get(key)
                if v is not None:
                    rows.append({'moc':moc,'tan_suat':'Ngày','chi_so_goc':label,'schema':key,'gia_tri':v,'nguon':'SBV liquidity scraper','file_nguon':'FA/data/sbv_liquidity/latest.json','note':'Nguồn SBV visible; daily job tự cập nhật.'})
        except Exception: pass
    return [r for r in rows if r.get('moc') and r.get('chi_so_goc')]

def map_indicator(r):
    text=keytext(r.get('chi_so_goc'), r.get('schema'), r.get('nguon'))
    for p in DROP_PATTERNS:
        if p.lower() in text:
            return None
    for cat, canon, includes, excludes in IMPORTANT_RULES:
        if any(i.lower() in text for i in includes) and not any(e.lower() in text for e in excludes):
            return cat, canon
    return None

def curated(rows):
    out=[]; dropped=[]
    for r in rows:
        m=map_indicator(r)
        if m:
            cat,canon=m; rr=dict(r); rr['nhom']=cat; rr['chi_so']=canon; out.append(rr)
        else:
            dropped.append(r)
    # de-duplicate same canonical/period/source: prefer official/SBV/Pinetree/TE over FiinProX when same period exists
    priority={'SBV liquidity scraper':1,'Pinetree Bản tin sáng':2,'TradingEconomics visible':3,'VCB':4,'WorldBank API':5,'FiinProX Excel/manual backfill':9}
    best={}
    for r in out:
        k=(r['nhom'],r['chi_so'],r['tan_suat'],r['moc'])
        cur=best.get(k)
        pr=priority.get(r['nguon'],5)
        if cur is None or pr < priority.get(cur['nguon'],5): best[k]=r
    return sorted(best.values(), key=lambda r:(r['nhom'],r['chi_so'],period_rank(r['moc']))), dropped

def pivot(rows, tan_suat=None):
    sub=[r for r in rows if tan_suat is None or r['tan_suat']==tan_suat]
    mocs=sorted(set(r['moc'] for r in sub), key=period_rank)
    meta={}; vals={}
    for r in sub:
        key=(r['nhom'],r['chi_so'])
        meta.setdefault(key, {'nguon':set(),'note':set(),'file':set(),'obs':0})
        meta[key]['nguon'].add(r['nguon']); meta[key]['note'].add(r['note']); meta[key]['file'].add(str(r.get('file_nguon') or '')); meta[key]['obs']+=1
        vals[(key,r['moc'])]=r['gia_tri']
    recs=[]
    for key,m in meta.items():
        nhom,chi=key
        rec={'nhom':nhom,'chi_so':chi,'nguon':'; '.join(sorted(m['nguon'])),'note':' | '.join(sorted(m['note'])),'file':'; '.join(sorted(x for x in m['file'] if x))[:700],'obs':m['obs']}
        for moc in mocs: rec[moc]=vals.get((key,moc),'')
        recs.append(rec)
    recs.sort(key=lambda r:(r['nhom'],r['chi_so']))
    return mocs,recs

def style(ws,mocs):
    blue=PatternFill('solid',fgColor='1F4E78'); yellow=PatternFill('solid',fgColor='FFF2CC'); green=PatternFill('solid',fgColor='E2F0D9')
    thin=Side(style='thin',color='D9D9D9')
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(wrap_text=True,vertical='top'); c.border=Border(bottom=thin)
    for c in ws[1]: c.fill=green; c.font=Font(bold=True)
    for c in ws[2]: c.fill=yellow; c.font=Font(italic=True)
    for c in ws[4]: c.fill=blue; c.font=Font(bold=True,color='FFFFFF')
    ws.freeze_panes='C5'; ws.auto_filter.ref=ws.dimensions
    ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=42
    for i in range(3,3+len(mocs)): ws.column_dimensions[get_column_letter(i)].width=13
    base=3+len(mocs)
    for j,w in enumerate([32,70,10,50]): ws.column_dimensions[get_column_letter(base+j)].width=w

def add_sheet(wb,name,mocs,recs,note):
    ws=wb.create_sheet(name[:31])
    ws.append(['NGUỒN','Bộ macro đã lọc: chỉ giữ chỉ số vĩ mô quan trọng cho Việt Nam, bỏ trùng ý nghĩa.'])
    ws.append(['GHI CHÚ',note])
    ws.append([])
    headers=['Nhóm','Chỉ số']+mocs+['Nguồn','Ghi chú','Số quan sát','File nguồn']
    ws.append(headers)
    for r in recs:
        ws.append([r['nhom'],r['chi_so']]+[r.get(m,'') for m in mocs]+[r['nguon'],r['note'],r['obs'],r['file']])
    style(ws,mocs)

def main():
    raw=collect_rows(); keep,drop=curated(raw)
    wb=Workbook(); ws=wb.active; ws.title='README'
    cats=defaultdict(int)
    for r in keep: cats[r['nhom']]+=1
    ws.append(['Bộ dữ liệu vĩ mô Việt Nam đã lọc - LH Investment'])
    ws.append(['Ngày tạo',datetime.now().isoformat(timespec='seconds')])
    ws.append(['Raw observations',len(raw)])
    ws.append(['Kept observations',len(keep)])
    ws.append(['Dropped observations',len(drop)])
    ws.append(['Nguyên tắc lọc','Giữ chỉ số đại diện cho regime Việt Nam: thanh khoản/lãi suất, tỷ giá-đối ngoại, tăng trưởng-lạm phát, thị trường-rủi ro, hàng hóa. Bỏ chỉ số trùng ý nghĩa hoặc ít cần thiết: nhiều cặp FX phụ, chỉ số chứng khoán nước ngoài phụ, HNX/VN30/UPCOM khi VNINDEX đã đại diện, M0/M1 nếu M2 đủ đại diện.'])
    ws.append([]); ws.append(['Nhóm','Số dòng giữ'])
    for k,v in sorted(cats.items()): ws.append([k,v])
    ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=120
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
    for ts,note in [
        ('Ngày','Dữ liệu theo ngày/trading-day; ưu tiên Pinetree/SBV/TradingEconomics visible.'),
        ('Tháng','Dữ liệu theo tháng; chủ yếu CPI, bán lẻ, IIP, FDI, M2/credit nếu có.'),
        ('Quý','Dữ liệu theo quý; chủ yếu GDP/BOP/cán cân nếu có.'),
        ('Năm','Dữ liệu năm; dùng làm bối cảnh dài hạn, không dùng làm trigger daily.'),
    ]:
        mocs,recs=pivot(keep,ts)
        if recs: add_sheet(wb,ts,mocs,recs,note)
    mocs,recs=pivot(keep,None); add_sheet(wb,'Tất cả đã lọc',mocs,recs,'Tất cả chỉ số quan trọng đã lọc, có thể dùng làm master macro input.')
    out=OUT/f'LH_Investment_Macro_Curated_Vietnam_VI_{datetime.now():%Y%m%d}.xlsx'
    wb.save(out)
    print(json.dumps({'out':str(out),'rawRows':len(raw),'keptRows':len(keep),'droppedRows':len(drop),'sheets':wb.sheetnames},ensure_ascii=False,indent=2))

if __name__=='__main__': main()
