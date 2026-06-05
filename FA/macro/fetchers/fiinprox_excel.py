"""
Fetcher/Importer: FiinProX Excel macro exports
==============================================

Purpose
- Convert manually downloaded FiinProX macro Excel exports under FA/ into one auditable long timeline.
- This is the production fallback for VN macro datasets that are hard/paywalled via public APIs:
  OMO history, SBV/statistical rates, bank deposit rates, M2/deposits/credit, BOP/trade balance, FDI.

Output row schema:
{
  "date": "YYYY-MM-DD|M-YYYY|Qx/YYYY",
  "indicator": "...",
  "value": 123.45,
  "source_file": "FiinProX_...xlsx",
  "sheet": "Sheet1",
  "category": "rates|fx|money_credit|trade_bop|liquidity_omo|macro_growth_inflation|other_macro",
  "source": "FiinProX Excel Export",
  "parserVersion": "fiinprox_excel_v1"
}
"""
from __future__ import annotations

import csv
import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

SOURCE_NAME = "FiinProX Excel Export"
PARSER_VERSION = "fiinprox_excel_v1"

FA_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUT_DIR = FA_ROOT / "data" / "unified_macro"

DATE_RE = re.compile(
    r"^(Q[1-4]/\d{4}|\d{1,2}[-/]\d{4}|\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}/\d{1,2}/\d{4})$"
)


def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip().replace("\n", " ")
    return re.sub(r"\s+", " ", s)


def _is_dateish(v: Any) -> bool:
    if isinstance(v, (datetime, date)):
        return True
    s = _norm(v)
    return bool(s and (DATE_RE.match(s) or re.match(r"^\d{1,2}-\d{4}$", s)))


def _clean_indicator(s: Any) -> str:
    s = _norm(s)
    s = re.sub(r"\s*Đơn vị:.*$", "", s, flags=re.I)
    return s.strip()


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def classify_indicator(name: str, file_name: str = "") -> str:
    low = (name + " " + file_name).lower()
    if any(x in low for x in ["omo", "nghiệp vụ thị trường mở", "thi truong mo", "tín phiếu", "tin phieu", "bơm", "bom", "hút", "hut"]):
        return "liquidity_omo"
    if any(x in low for x in ["lãi suất", "lai suat", "overnight", "qua đêm", "tái cấp vốn", "tái chiết khấu", "huy động", "lshđ"]):
        return "rates"
    if any(x in low for x in ["usd", "eur", "cny", "tỷ giá", "ty gia", "fx", "vnd-usd"]):
        return "fx"
    if any(x in low for x in ["m2", "tiền gửi", "tin dung", "tín dụng", "dư nợ"]):
        return "money_credit"
    if any(x in low for x in ["xuất khẩu", "xuat khau", "nhập khẩu", "nhap khau", "cán cân", "can can", "f.o.b", "vãng lai"]):
        return "trade_bop"
    if any(x in low for x in ["gdp", "cpi", "iip", "pmi", "fdi", "bán lẻ", "ban le", "retail"]):
        return "macro_growth_inflation"
    return "other_macro"


def extract_file(path: Path) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")
    rows: list[dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        data = [list(r) for r in ws.iter_rows(values_only=True)]
        if not data:
            continue

        # Pattern A: long-ish table. Header row has Ngày/date; metric columns follow.
        for hi, row in enumerate(data[:15]):
            vals = [_norm(x) for x in row]
            if any(v.lower() in ("ngày", "ngay", "date") for v in vals) and len([v for v in vals if v]) >= 4:
                date_idx = next(i for i, v in enumerate(vals) if v.lower() in ("ngày", "ngay", "date"))
                headers = vals
                for rr in data[hi + 1:]:
                    d = _norm(rr[date_idx] if date_idx < len(rr) else "")
                    if not d or not _is_dateish(d):
                        continue
                    for ci, h in enumerate(headers):
                        if ci <= date_idx or not h:
                            continue
                        fv = _to_float(rr[ci] if ci < len(rr) else None)
                        if fv is None:
                            continue
                        ind = _clean_indicator(h)
                        rows.append(_row(d, ind, fv, path, ws.title, "row_table"))
                break

        # Pattern B: wide table. Header row has many date/period columns, first col is indicator.
        for hi, row in enumerate(data[:12]):
            headers = [_norm(x) for x in row]
            date_cols = [i for i, h in enumerate(headers) if _is_dateish(h)]
            if len(date_cols) >= 2:
                for rr in data[hi + 1:]:
                    ind = _clean_indicator(rr[0] if rr else "")
                    if not ind or ind.lower() in ("stt", "kỳ", "ngày", "date"):
                        continue
                    for ci in date_cols:
                        fv = _to_float(rr[ci] if ci < len(rr) else None)
                        if fv is None:
                            continue
                        rows.append(_row(headers[ci], ind, fv, path, ws.title, "wide_table"))
                break
    return rows


def _row(d: str, ind: str, val: float, path: Path, sheet: str, orientation: str) -> dict[str, Any]:
    return {
        "date": d,
        "category": classify_indicator(ind, path.name),
        "indicator": ind,
        "value": val,
        "source_file": path.name,
        "sheet": sheet,
        "orientation": orientation,
        "source": SOURCE_NAME,
        "parserVersion": PARSER_VERSION,
    }


def fetch(folder: str | Path | None = None) -> dict[str, Any]:
    folder = Path(folder) if folder else FA_ROOT
    files = sorted(folder.glob("FiinProX_*.xlsx"))
    rows: list[dict[str, Any]] = []
    errors: dict[str, str] = {}
    for f in files:
        try:
            rows.extend(extract_file(f))
        except Exception as e:
            errors[f.name] = str(e)[:200]

    seen = set()
    unique = []
    for r in rows:
        key = (r["date"], r["indicator"], r["value"], r["source_file"], r["sheet"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)
    unique.sort(key=lambda r: (str(r["date"]), r["category"], r["source_file"], r["indicator"]))
    return {
        "source": SOURCE_NAME,
        "parserVersion": PARSER_VERSION,
        "fetchedAt": datetime.now().isoformat(),
        "status": "ok" if unique else "no_rows",
        "fileCount": len(files),
        "rowCount": len(unique),
        "indicatorCount": len({r["indicator"] for r in unique}),
        "rows": unique,
        "errors": errors or None,
    }


def save_unified(result: dict[str, Any], out_dir: str | Path | None = None) -> dict[str, str]:
    out = Path(out_dir) if out_dir else DEFAULT_OUT_DIR
    out.mkdir(parents=True, exist_ok=True)
    rows = result.get("rows", [])
    csv_path = out / "macro_timeline_unified.csv"
    json_path = out / "macro_timeline_unified.json"
    fields = ["date", "category", "indicator", "value", "source_file", "sheet", "orientation", "source", "parserVersion"]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in fields})
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"csv": str(csv_path), "json": str(json_path)}


if __name__ == "__main__":
    res = fetch()
    paths = save_unified(res)
    print(json.dumps({"rowCount": res["rowCount"], "indicatorCount": res["indicatorCount"], **paths}, ensure_ascii=False, indent=2))
