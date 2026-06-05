from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "reports"
OUT.mkdir(exist_ok=True)


def load_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists(): return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v: Any) -> float | None:
    if v is None or v == "": return None
    try: return float(str(v).replace(",", ""))
    except Exception: return None


def year_of(period: str) -> int | None:
    s = str(period or "")
    m = re.search(r"(20\d{2}|19\d{2})", s)
    return int(m.group(1)) if m else None


def period_rank(period: str) -> tuple:
    s = str(period or "")
    # Larger rank means later inside the year.
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m: return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"^(\d{1,2})-(\d{4})$", s)
    if m: return (int(m.group(2)), int(m.group(1)), 28)
    m = re.match(r"^Q([1-4])/(\d{4})$", s, re.I)
    if m: return (int(m.group(2)), int(m.group(1))*3, 30)
    y = year_of(s)
    return (y or 0, 12, 31)


def classify_period(period: str, source_file: str = "", frequency: str = "") -> str:
    s = str(period or "")
    sf = str(source_file or "").lower()
    f = str(frequency or "").lower()
    if "can can" in sf or "cán cân" in sf or re.match(r"^Q[1-4]/\d{4}$", s, re.I): return "Quarterly"
    if any(x in sf for x in ["du_lieu_vi_mo", "lai suat huy dong", "lai suat thong ke", "20266"]): return "Monthly"
    if "annual" in f or re.match(r"^\d{4}$", s): return "Annual"
    if re.match(r"^\d{1,2}-\d{4}$", s): return "Monthly"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s): return "Daily"
    return "Mixed"


def norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def collect_rows() -> list[dict[str, Any]]:
    rows=[]
    # Pinetree archive
    for r in load_csv(ROOT/'data/pinetree_archive/pinetree_macro_timeline.csv'):
        rows.append({
            'period': r.get('date'), 'year': year_of(r.get('date')), 'periodType': 'Daily',
            'indicator': norm(r.get('label') or r.get('indicator')), 'value': num(r.get('value')),
            'source': 'Pinetree Morning Brief archive', 'sourceFile': 'FA/data/pinetree_archive/pinetree_macro_timeline.csv',
            'url': r.get('url'), 'rank': period_rank(r.get('date')),
            'note': 'Daily/trading-day. Public Pinetree Bản tin sáng; can update daily automatically. Year cell uses latest available observation in that year.'
        })
    # FiinProX unified
    for r in load_csv(ROOT/'data/unified_macro/macro_timeline_unified.csv'):
        pt=classify_period(r.get('date'), r.get('source_file'))
        rows.append({
            'period': r.get('date'), 'year': year_of(r.get('date')), 'periodType': pt,
            'indicator': norm(r.get('indicator')), 'value': num(r.get('value')),
            'source': 'FiinProX Excel export', 'sourceFile': r.get('source_file'), 'url': '', 'rank': period_rank(r.get('date')),
            'note': 'Manual/premium export. Not guaranteed daily; update when new FiinProX Excel is placed in FA/. Year cell uses latest period in that year.'
        })
    # TradingEconomics visible
    for r in load_csv(ROOT/'data/tradingeconomics_visible_timeline.csv'):
        period=(r.get('fetchedAt') or '')[:10]
        rows.append({
            'period': period, 'year': year_of(period), 'periodType': 'Daily',
            'indicator': norm(r.get('indicator')), 'value': num(r.get('value')),
            'source': 'TradingEconomics visible browser scrape', 'sourceFile': 'FA/data/tradingeconomics_visible_timeline.csv',
            'url': r.get('url'), 'rank': period_rank(period),
            'note': 'Daily/latest visible public page only via headed Chrome; no Download/API/subscription bypass. Year cell uses latest scrape in that year.'
        })
    # WorldBank from latest history with WB
    hist=sorted((ROOT/'data/history').glob('2026-06-05_v*.json'))
    if hist:
        for chosen in reversed(hist):
            try:
                snap=json.loads(chosen.read_text(encoding='utf-8'))
                wb=snap.get('worldbank',{}).get('data',{})
                if wb: break
            except Exception: wb={}
        for key,obj in wb.items():
            for yr,val in (obj.get('timeSeries') or {}).items():
                if val is None: continue
                rows.append({
                    'period': yr, 'year': year_of(yr), 'periodType': 'Annual',
                    'indicator': norm(key), 'value': num(val), 'source': 'WorldBank API',
                    'sourceFile': str(chosen.relative_to(ROOT)), 'url': 'https://api.worldbank.org/', 'rank': period_rank(yr),
                    'note': 'Annual official WorldBank data; refreshed automatically but underlying value is lagged. Year cell is annual value.'
                })
    return [r for r in rows if r.get('year') and r.get('indicator') and r.get('value') is not None]


def aggregate(rows: list[dict[str, Any]], period_filter: str | None = None):
    if period_filter:
        rows=[r for r in rows if r['periodType']==period_filter]
    years=sorted(set(r['year'] for r in rows))
    bucket={}
    meta={}
    for r in rows:
        key=(r['indicator'], r['source'], r['periodType'])
        yr=r['year']
        cur=bucket.get((key,yr))
        if cur is None or r['rank']>cur['rank']:
            bucket[(key,yr)]=r
        m=meta.setdefault(key, {'sourceFiles':set(), 'urls':set(), 'notes':set(), 'count':0})
        if r.get('sourceFile'): m['sourceFiles'].add(str(r['sourceFile']))
        if r.get('url'): m['urls'].add(str(r['url']))
        if r.get('note'): m['notes'].add(str(r['note']))
        m['count']+=1
    records=[]
    for key,m in meta.items():
        indicator,source,pt=key
        rec={'periodType':pt,'indicator':indicator,'source':source,'sourceFiles':'; '.join(sorted(m['sourceFiles']))[:800], 'note':' | '.join(sorted(m['notes']))[:1000], 'observations':m['count']}
        for y in years:
            rec[y]=bucket.get((key,y),{}).get('value','')
        records.append(rec)
    records.sort(key=lambda r:(r['periodType'], r['source'], r['indicator']))
    return years,records


def style(ws, year_cols_count):
    blue=PatternFill('solid', fgColor='1F4E78')
    yellow=PatternFill('solid', fgColor='FFF2CC')
    sourcefill=PatternFill('solid', fgColor='D9EAF7')
    thin=Side(style='thin', color='D9D9D9')
    for row in ws.iter_rows():
        for c in row:
            c.alignment=Alignment(wrap_text=True, vertical='top')
            c.border=Border(bottom=thin)
    for c in ws[1]: c.fill=sourcefill; c.font=Font(bold=True)
    for c in ws[2]: c.fill=yellow; c.font=Font(italic=True)
    for c in ws[4]: c.fill=blue; c.font=Font(bold=True,color='FFFFFF')
    ws.freeze_panes='B5'
    ws.auto_filter.ref=ws.dimensions
    ws.column_dimensions['A'].width=16
    ws.column_dimensions['B'].width=46
    for i in range(3,3+year_cols_count): ws.column_dimensions[get_column_letter(i)].width=13
    ws.column_dimensions[get_column_letter(3+year_cols_count)].width=30
    ws.column_dimensions[get_column_letter(4+year_cols_count)].width=55
    ws.column_dimensions[get_column_letter(5+year_cols_count)].width=12


def write_sheet(wb,name,years,records,source_line,note_line):
    ws=wb.create_sheet(name[:31])
    ws.append(['SOURCE', source_line])
    ws.append(['NOTE', note_line])
    ws.append([])
    headers=['periodType','indicator']+[str(y) for y in years]+['source','note','observations','sourceFiles']
    ws.append(headers)
    for r in records:
        ws.append([r.get('periodType'), r.get('indicator')]+[r.get(y,'') for y in years]+[r.get('source'), r.get('note'), r.get('observations'), r.get('sourceFiles')])
    style(ws,len(years))


def main():
    rows=collect_rows()
    wb=Workbook()
    ws=wb.active; ws.title='README'
    years_all, records_all=aggregate(rows)
    ws.append(['LH Investment Macro Year Matrix'])
    ws.append(['Created', datetime.now().isoformat(timespec='seconds')])
    ws.append(['Input observations', len(rows)])
    ws.append(['Matrix rows', len(records_all)])
    ws.append(['Rule', 'Mỗi ô năm lấy observation mới nhất trong năm đó cho từng chỉ số + nguồn + periodType. Cuối mỗi hàng có source, note, số observations, sourceFiles.'])
    ws.append([])
    ws.append(['Sheet','Meaning'])
    for name,meaning in [('All_Year_Matrix','Tất cả chỉ số'),('Daily','Chỉ số ngày/trading-day'),('Monthly','Chỉ số tháng'),('Quarterly','Chỉ số quý'),('Annual','Chỉ số năm')]: ws.append([name,meaning])
    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=120
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(wrap_text=True, vertical='top')

    write_sheet(wb,'All_Year_Matrix',years_all,records_all,'All macro sources currently available','Rows are indicator x source x periodType; columns are years; value is latest available observation inside each year.')
    for pt in ['Daily','Monthly','Quarterly','Annual']:
        years,recs=aggregate(rows,pt)
        if recs:
            note={
                'Daily':'Daily/trading-day data. Can update automatically for Pinetree/TradingEconomics visible/Yahoo/VCB/SBV daily sources; FiinProX daily rows need manual export.',
                'Monthly':'Monthly data. Mostly FiinProX/manual exports; public GSO monthly source still unavailable from this environment.',
                'Quarterly':'Quarterly data. Mostly FiinProX BOP/trade/capital account detail.',
                'Annual':'Annual slow macro context. WorldBank API auto-refreshes but actual values are lagged.'
            }[pt]
            write_sheet(wb,pt,years,recs,pt+' macro sources',note)
    out=OUT/f'LH_Investment_Macro_Year_Matrix_{datetime.now():%Y%m%d}.xlsx'
    wb.save(out)
    print(json.dumps({'out':str(out),'inputRows':len(rows),'matrixRows':len(records_all),'sheets':wb.sheetnames},ensure_ascii=False,indent=2))

if __name__=='__main__': main()
