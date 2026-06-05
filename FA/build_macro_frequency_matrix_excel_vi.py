from __future__ import annotations

import csv, json, re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parent
OUT=ROOT/'reports'; OUT.mkdir(exist_ok=True)


def load_csv(p:Path):
    if not p.exists(): return []
    with p.open('r',encoding='utf-8-sig',newline='') as f: return list(csv.DictReader(f))

def num(v):
    if v is None or v=='': return None
    try: return float(str(v).replace(',',''))
    except: return None

def norm(s): return re.sub(r'\s+',' ',str(s or '').strip())

def period_rank(s):
    s=str(s or '')
    m=re.match(r'^(\d{4})-(\d{2})-(\d{2})$',s)
    if m: return (int(m.group(1)),int(m.group(2)),int(m.group(3)))
    m=re.match(r'^(\d{1,2})-(\d{4})$',s)
    if m: return (int(m.group(2)),int(m.group(1)),28)
    m=re.match(r'^Q([1-4])/(\d{4})$',s,re.I)
    if m: return (int(m.group(2)),int(m.group(1))*3,30)
    m=re.match(r'^(\d{4})$',s)
    if m: return (int(m.group(1)),12,31)
    return (0,0,0)

def classify(period, source_file='', frequency=''):
    s=str(period or ''); sf=str(source_file or '').lower(); f=str(frequency or '').lower()
    if 'can can' in sf or 'cán cân' in sf or re.match(r'^Q[1-4]/\d{4}$',s,re.I): return 'Quý'
    if any(x in sf for x in ['du_lieu_vi_mo','lai suat huy dong','lai suat thong ke','20266']): return 'Tháng'
    if re.match(r'^\d{1,2}-\d{4}$',s) or 'monthly' in f: return 'Tháng'
    if 'nghiep vu thi truong mo' in sf or 'daily' in f or re.match(r'^\d{4}-\d{2}-\d{2}$',s): return 'Ngày'
    if re.match(r'^\d{4}$',s) or 'annual' in f: return 'Năm'
    return 'Khác'

def collect():
    rows=[]
    # Pinetree daily archive
    for r in load_csv(ROOT/'data/pinetree_archive/pinetree_macro_timeline.csv'):
        v=num(r.get('value'))
        if v is None: continue
        rows.append({'tan_suat':'Ngày','moc':r.get('date'),'chi_so':norm(r.get('label') or r.get('indicator')),'gia_tri':v,'nguon':'Pinetree - Bản tin sáng','file_nguon':'FA/data/pinetree_archive/pinetree_macro_timeline.csv','url':r.get('url'),'ghi_chu':'Cập nhật được hằng ngày theo ngày giao dịch. Dữ liệu lấy từ archive Bản tin sáng Pinetree; ô là giá trị tại đúng ngày đó.'})
    # FiinProX
    for r in load_csv(ROOT/'data/unified_macro/macro_timeline_unified.csv'):
        v=num(r.get('value'))
        if v is None: continue
        ts=classify(r.get('date'),r.get('source_file'))
        daily='Cần file export mới từ FiinProX; không phải nguồn tự động hằng ngày.'
        if ts=='Ngày': daily='Dữ liệu dạng ngày trong FiinProX; chỉ cập nhật khi có file export mới.'
        rows.append({'tan_suat':ts,'moc':r.get('date'),'chi_so':norm(r.get('indicator')),'gia_tri':v,'nguon':'FiinProX Excel','file_nguon':r.get('source_file'),'url':'','ghi_chu':daily})
    # TradingEconomics visible latest
    for r in load_csv(ROOT/'data/tradingeconomics_visible_timeline.csv'):
        v=num(r.get('value'))
        if v is None: continue
        moc=(r.get('fetchedAt') or '')[:10]
        rows.append({'tan_suat':'Ngày','moc':moc,'chi_so':norm(r.get('indicator')),'gia_tri':v,'nguon':'TradingEconomics - scrape trang public','file_nguon':'FA/data/tradingeconomics_visible_timeline.csv','url':r.get('url'),'ghi_chu':'Cập nhật được hằng ngày nhưng chỉ là dữ liệu visible/latest trên trang public; không tải CSV/API.'})
    # WorldBank annual
    wb={}; chosen=None
    for p in reversed(sorted((ROOT/'data/history').glob('2026-06-05_v*.json'))):
        try:
            snap=json.loads(p.read_text(encoding='utf-8')); wb=snap.get('worldbank',{}).get('data',{})
            if wb: chosen=p; break
        except: pass
    for key,obj in wb.items():
        for yr,v in (obj.get('timeSeries') or {}).items():
            if v is None: continue
            rows.append({'tan_suat':'Năm','moc':yr,'chi_so':norm(key),'gia_tri':v,'nguon':'WorldBank API','file_nguon':str(chosen.relative_to(ROOT)) if chosen else '','url':'https://api.worldbank.org/','ghi_chu':'Dữ liệu năm; runner có thể gọi hằng ngày nhưng số liệu thực tế cập nhật chậm/trễ nhiều tháng.'})
    return [r for r in rows if r['moc'] and r['chi_so']]

def pivot(rows, tan_suat):
    sub=[r for r in rows if r['tan_suat']==tan_suat]
    mocs=sorted(set(r['moc'] for r in sub), key=period_rank)
    bucket={}; meta={}
    for r in sub:
        key=(r['chi_so'],r['nguon'])
        bucket[(key,r['moc'])]=r['gia_tri']
        mt=meta.setdefault(key, {'ghi_chu':set(),'file_nguon':set(),'url':set(),'obs':0})
        mt['ghi_chu'].add(r['ghi_chu']); mt['obs']+=1
        if r.get('file_nguon'): mt['file_nguon'].add(str(r['file_nguon']))
        if r.get('url'): mt['url'].add(str(r['url']))
    recs=[]
    for key,mt in meta.items():
        chi_so,nguon=key
        rec={'chi_so':chi_so,'nguon':nguon,'so_quan_sat':mt['obs'],'ghi_chu':' | '.join(sorted(mt['ghi_chu'])),'file_nguon':'; '.join(sorted(mt['file_nguon']))[:800], 'url':'; '.join(sorted(mt['url']))[:800]}
        for m in mocs: rec[m]=bucket.get((key,m),'')
        recs.append(rec)
    recs.sort(key=lambda x:(x['nguon'],x['chi_so']))
    return mocs,recs

def style(ws, mocs):
    blue=PatternFill('solid',fgColor='1F4E78'); yellow=PatternFill('solid',fgColor='FFF2CC'); src=PatternFill('solid',fgColor='D9EAF7')
    thin=Side(style='thin',color='D9D9D9')
    for row in ws.iter_rows():
        for c in row:
            c.alignment=Alignment(wrap_text=True,vertical='top'); c.border=Border(bottom=thin)
    for c in ws[1]: c.fill=src; c.font=Font(bold=True)
    for c in ws[2]: c.fill=yellow; c.font=Font(italic=True)
    for c in ws[4]: c.fill=blue; c.font=Font(bold=True,color='FFFFFF')
    ws.freeze_panes='B5'; ws.auto_filter.ref=ws.dimensions
    ws.column_dimensions['A'].width=48
    for i in range(2,2+len(mocs)): ws.column_dimensions[get_column_letter(i)].width=13
    base=2+len(mocs)
    ws.column_dimensions[get_column_letter(base)].width=28
    ws.column_dimensions[get_column_letter(base+1)].width=70
    ws.column_dimensions[get_column_letter(base+2)].width=12
    ws.column_dimensions[get_column_letter(base+3)].width=55

def add_sheet(wb,name,mocs,recs,note):
    ws=wb.create_sheet(name)
    ws.append(['NGUỒN', 'Tổng hợp từ Pinetree, FiinProX, TradingEconomics visible, WorldBank'])
    ws.append(['GHI CHÚ', note])
    ws.append([])
    headers=['Chỉ số']+mocs+['Nguồn','Ghi chú','Số quan sát','File nguồn']
    ws.append(headers)
    for r in recs:
        ws.append([r['chi_so']]+[r.get(m,'') for m in mocs]+[r['nguon'],r['ghi_chu'],r['so_quan_sat'],r['file_nguon']])
    style(ws,mocs)

def main():
    rows=collect()
    wb=Workbook(); ws=wb.active; ws.title='README'
    ws.append(['Bảng tổng hợp dữ liệu vĩ mô LH Investment'])
    ws.append(['Ngày tạo', datetime.now().isoformat(timespec='seconds')])
    ws.append(['Số dòng gốc', len(rows)])
    ws.append(['Cách trình bày', 'Hàng ngang là mốc thời gian đúng tần suất: ngày/tháng/quý/năm. Hàng dọc là chỉ số. Cuối mỗi hàng có Nguồn, Ghi chú, Số quan sát, File nguồn.'])
    ws.append(['Lưu ý', 'Sheet Ngày giữ từng ngày; không gom theo năm. Sheet Tháng giữ từng tháng; Sheet Quý giữ từng quý; Sheet Năm giữ từng năm.'])
    ws.column_dimensions['A'].width=24; ws.column_dimensions['B'].width=120
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
    sheet_notes={
        'Ngày':'Dữ liệu theo từng ngày/trading-day. Pinetree và TradingEconomics visible có thể cập nhật hằng ngày tự động; FiinProX dạng ngày cần export mới.',
        'Tháng':'Dữ liệu theo từng tháng. Hiện chủ yếu từ FiinProX Excel; GSO public chưa truy cập được ổn định từ môi trường này.',
        'Quý':'Dữ liệu theo từng quý, chủ yếu cán cân thanh toán/thương mại từ FiinProX.',
        'Năm':'Dữ liệu theo từng năm, chủ yếu từ WorldBank API và một số FiinProX/annual context.'
    }
    for ts in ['Ngày','Tháng','Quý','Năm']:
        mocs,recs=pivot(rows,ts)
        if recs: add_sheet(wb,ts,mocs,recs,sheet_notes[ts])
    out=OUT/f'LH_Investment_Macro_Frequency_Matrix_VI_{datetime.now():%Y%m%d}.xlsx'
    wb.save(out)
    print(json.dumps({'out':str(out),'rows':len(rows),'sheets':wb.sheetnames},ensure_ascii=False,indent=2))
if __name__=='__main__': main()
