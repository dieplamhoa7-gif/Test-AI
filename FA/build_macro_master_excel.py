from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "reports"
OUT.mkdir(exist_ok=True)


def load_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def norm_indicator(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def parse_num(v: Any) -> float | None:
    if v is None or v == "": return None
    try: return float(str(v).replace(",", ""))
    except Exception: return None


def period_type(d: str, source_file: str = "", frequency: str = "") -> str:
    s = str(d or "").strip()
    f = (frequency or "").lower()
    sf = (source_file or "").lower()
    # Source-file hints first because many monthly FiinProX rows use month-end ISO dates.
    if "can can" in sf or "cán cân" in sf or "quarter" in f:
        return "Quarterly"
    if any(x in sf for x in ["du_lieu_vi_mo", "lai suat huy dong", "lai suat thong ke", "20266"]):
        return "Monthly"
    if "nghiep vu thi truong mo" in sf or "daily" in f:
        return "Daily"
    if re.match(r"^Q[1-4]/\d{4}$", s, re.I):
        return "Quarterly"
    if re.match(r"^\d{1,2}-\d{4}$", s) or "monthly" in f:
        return "Monthly"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return "Daily"
    if re.match(r"^\d{4}$", s) or "annual" in f:
        return "Annual"
    return "Mixed"


def add_rows_from_pinetree(rows: list[dict[str, Any]]):
    src = ROOT / "data" / "pinetree_archive" / "pinetree_macro_timeline.csv"
    for r in load_csv(src):
        rows.append({
            "period": r.get("date"),
            "periodType": "Daily",
            "indicator": norm_indicator(r.get("label") or r.get("indicator")),
            "schemaKey": r.get("indicator"),
            "value": parse_num(r.get("value")),
            "change1d": parse_num(r.get("change1d")),
            "ytd": parse_num(r.get("ytd")),
            "source": "Pinetree Morning Brief archive",
            "sourceFile": "FA/data/pinetree_archive/pinetree_macro_timeline.csv",
            "url": r.get("url"),
            "dailyUpdatable": "YES",
            "note": "Public Pinetree Bản tin sáng; daily/trading-day update; already archived from category pages.",
        })


def add_rows_from_fiin(rows: list[dict[str, Any]]):
    src = ROOT / "data" / "unified_macro" / "macro_timeline_unified.csv"
    for r in load_csv(src):
        pt = period_type(r.get("date"), r.get("source_file"), "")
        daily = "NO"
        note = "FiinProX manual/premium fallback; update only when a new Excel export is placed in FA/."
        if pt == "Daily":
            daily = "MANUAL_DAILY_IF_EXPORT_EXISTS"
        rows.append({
            "period": r.get("date"),
            "periodType": pt,
            "indicator": norm_indicator(r.get("indicator")),
            "schemaKey": "",
            "value": parse_num(r.get("value")),
            "change1d": "",
            "ytd": "",
            "source": "FiinProX Excel export",
            "sourceFile": r.get("source_file"),
            "url": "",
            "dailyUpdatable": daily,
            "note": note,
        })


def add_rows_from_te(rows: list[dict[str, Any]]):
    src = ROOT / "data" / "tradingeconomics_visible_timeline.csv"
    for r in load_csv(src):
        rows.append({
            "period": (r.get("fetchedAt") or "")[:10],
            "periodType": "Daily",
            "indicator": norm_indicator(r.get("indicator")),
            "schemaKey": r.get("key"),
            "value": parse_num(r.get("value")),
            "change1d": "",
            "ytd": "",
            "source": "TradingEconomics visible browser scrape",
            "sourceFile": "FA/data/tradingeconomics_visible_timeline.csv",
            "url": r.get("url"),
            "dailyUpdatable": "YES_VISIBLE_ONLY",
            "note": "Scraped visible public page data via headed Chrome; no Download/API/subscription bypass; latest visible values only.",
        })


def add_rows_from_worldbank(rows: list[dict[str, Any]]):
    hist = sorted((ROOT / "data" / "history").glob("2026-06-05_v*.json"))
    chosen = hist[-1] if hist else None
    if not chosen or not chosen.exists(): return
    try:
        snap = json.loads(chosen.read_text(encoding="utf-8"))
    except Exception:
        return
    wb = snap.get("worldbank", {}).get("data", {})
    for key, obj in wb.items():
        for yr, val in (obj.get("timeSeries") or {}).items():
            if val is None: continue
            rows.append({
                "period": yr,
                "periodType": "Annual",
                "indicator": key,
                "schemaKey": key,
                "value": val,
                "change1d": "",
                "ytd": "",
                "source": "WorldBank API",
                "sourceFile": str(chosen.relative_to(ROOT)),
                "url": "https://api.worldbank.org/",
                "dailyUpdatable": "NO_ANNUAL_LAGGED",
                "note": "Official WorldBank annual indicator; refreshed daily if runner runs, but actual data changes slowly and lags months.",
            })


def summarize_sources(rows: list[dict[str, Any]]):
    by = defaultdict(lambda: {"rows": 0, "periodTypes": set(), "indicators": set(), "dailyFlags": set(), "notes": set()})
    for r in rows:
        b = by[r["source"]]
        b["rows"] += 1
        b["periodTypes"].add(r.get("periodType"))
        b["indicators"].add(r.get("indicator"))
        b["dailyFlags"].add(r.get("dailyUpdatable"))
        if r.get("note"): b["notes"].add(r["note"])
    out = []
    for src, b in sorted(by.items()):
        out.append({
            "source": src,
            "rows": b["rows"],
            "periodTypes": ", ".join(sorted(x for x in b["periodTypes"] if x)),
            "indicatorCount": len(b["indicators"]),
            "dailyUpdatable": ", ".join(sorted(x for x in b["dailyFlags"] if x)),
            "note": " | ".join(sorted(b["notes"]))[:600],
        })
    return out


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    source_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = source_fill; cell.font = Font(bold=True)
    for cell in ws[2]:
        cell.fill = note_fill; cell.font = Font(italic=True)
    for cell in ws[4]:
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = min(60, max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200) + 1)) + 2)
        ws.column_dimensions[letter].width = max(12, max_len)


def write_data_sheet(wb, name: str, rows: list[dict[str, Any]], source_line: str, note_line: str):
    ws = wb.create_sheet(name[:31])
    headers = ["period", "periodType", "indicator", "schemaKey", "value", "change1d", "ytd", "source", "sourceFile", "dailyUpdatable", "note", "url"]
    ws.append(["SOURCE", source_line])
    ws.append(["NOTE", note_line])
    ws.append([])
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    style_sheet(ws)


def main():
    rows: list[dict[str, Any]] = []
    add_rows_from_pinetree(rows)
    add_rows_from_fiin(rows)
    add_rows_from_te(rows)
    add_rows_from_worldbank(rows)

    rows = [r for r in rows if r.get("period") and r.get("indicator") and r.get("value") not in (None, "")]
    rows.sort(key=lambda r: (r["periodType"], str(r["period"]), r["source"], r["indicator"]))

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "README"
    ws0.append(["LH Investment Macro Master Data"])
    ws0.append(["Created", datetime.now().isoformat(timespec="seconds")])
    ws0.append(["Total rows", len(rows)])
    ws0.append(["Purpose", "Tổng hợp tất cả data vĩ mô hiện có theo khung ngày/tháng/quý/năm. Mỗi data sheet có dòng SOURCE và NOTE ở 2 dòng đầu."])
    ws0.append(["Daily update note", "YES = có thể cập nhật hằng ngày tự động; YES_VISIBLE_ONLY = chỉ latest visible page; MANUAL_DAILY_IF_EXPORT_EXISTS = cần file Excel export mới; NO_ANNUAL_LAGGED = dữ liệu năm, cập nhật chậm."])
    ws0.append([])
    ws0.append(["Source", "Rows", "Period types", "Indicator count", "Daily updatable", "Note"])
    for s in summarize_sources(rows):
        ws0.append([s["source"], s["rows"], s["periodTypes"], s["indicatorCount"], s["dailyUpdatable"], s["note"]])
    for c in ws0[1]: c.font = Font(bold=True, size=14)
    ws0.column_dimensions["A"].width = 38
    ws0.column_dimensions["B"].width = 12
    ws0.column_dimensions["C"].width = 25
    ws0.column_dimensions["D"].width = 16
    ws0.column_dimensions["E"].width = 28
    ws0.column_dimensions["F"].width = 80
    for row in ws0.iter_rows():
        for cell in row: cell.alignment = Alignment(wrap_text=True, vertical="top")

    for pt in ["Daily", "Monthly", "Quarterly", "Annual", "Mixed"]:
        subset = [r for r in rows if r["periodType"] == pt]
        if not subset: continue
        sources = sorted(set(r["source"] for r in subset))
        source_line = "; ".join(sources)
        note_line = {
            "Daily": "Daily/trading-day data. Pinetree, VCB/Yahoo/TradingEconomics visible can update daily; FiinProX daily rows need new export if any.",
            "Monthly": "Monthly macro series. Mostly FiinProX/manual export currently; WorldBank not monthly; GSO still unavailable from this environment.",
            "Quarterly": "Quarterly BOP/trade/capital account data. Mostly FiinProX/manual export; public automatic source not fully available yet.",
            "Annual": "Annual slow macro context from WorldBank API and/or FiinProX; refreshed daily but actual values lag months.",
            "Mixed": "Rows with unclear/other period format; verify before model use.",
        }[pt]
        write_data_sheet(wb, pt, subset, source_line, note_line)

    out = OUT / f"LH_Investment_Macro_Master_Data_{datetime.now():%Y%m%d}.xlsx"
    wb.save(out)
    print(json.dumps({"out": str(out), "rows": len(rows), "sheets": wb.sheetnames}, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
