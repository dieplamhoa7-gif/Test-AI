from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.comments import Comment
from pathlib import Path

out = Path('MWG_valuation_model_basic.xlsx')
wb = Workbook()
ws = wb.active
ws.title = 'README'

blue = PatternFill('solid', fgColor='1F4E78')
light = PatternFill('solid', fgColor='D9EAF7')
yellow = PatternFill('solid', fgColor='FFF2CC')
green = PatternFill('solid', fgColor='E2F0D9')
red = PatternFill('solid', fgColor='FCE4D6')
gray = PatternFill('solid', fgColor='E7E6E6')
white_font = Font(color='FFFFFF', bold=True)
bold = Font(bold=True)
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col in range(1, 12): ws.column_dimensions[get_column_letter(col)].width = 18
ws['A1'] = 'MWG BASIC VALUATION MODEL - DCF + SOTP + IPO ĐMX + BHX'
ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws['A3'] = 'Mục tiêu'
ws['B3'] = 'Khung Excel định giá MWG bản cơ bản, có thể chỉnh assumptions. Dữ liệu lịch sử là placeholder/assumption cần cập nhật bằng BCTC chính thức.'
ws['A5'] = 'Sheets'
rows = [
('Assumptions','Các giả định chính'),('Historical_FS','BCTC lịch sử/normalized'),('DCF_Consolidated','DCF hợp nhất'),('SOTP','Sum-of-the-parts TGDD/ĐMX/BHX/khác'),('DMX_IPO','Phân tích IPO ĐMX và tác động đến MWG'),('BHX_Analysis','Vận hành BHX, thị phần, triển vọng'),('Macro_VN','Macro Việt Nam'),('Sensitivity','Độ nhạy WACC/g'),('Sources','Nguồn cần cập nhật')]
for i,r in enumerate(rows,6): ws.cell(i,1).value, ws.cell(i,2).value = r
ws['A17'] = 'Lưu ý quan trọng'
ws['B17'] = 'Không phải khuyến nghị đầu tư. Cần cập nhật BCTC, giá cổ phiếu, số cổ phiếu, nợ tiền mặt từ nguồn chính thức trước khi ra quyết định.'

# Assumptions
ws = wb.create_sheet('Assumptions')
for col in range(1, 8): ws.column_dimensions[get_column_letter(col)].width = 22
headers = ['Input','Base','Bull','Bear','Unit','Ghi chú','Editable']
ws.append(headers)
for c in range(1,8): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
assumptions = [
('Revenue 2024', 134000, 134000, 134000, 'VNDbn', 'Placeholder; cập nhật theo BCTC MWG', 'Y'),
('Revenue growth 2025', 0.11, 0.15, 0.06, '%', 'Phục hồi ICT/CE + BHX tăng trưởng', 'Y'),
('Revenue growth 2026', 0.09, 0.12, 0.05, '%', '', 'Y'),
('Revenue growth 2027', 0.07, 0.10, 0.04, '%', '', 'Y'),
('Revenue growth 2028', 0.05, 0.08, 0.03, '%', '', 'Y'),
('Revenue growth 2029', 0.04, 0.06, 0.02, '%', '', 'Y'),
('EBIT margin 2024', 0.035, 0.035, 0.035, '%', 'Normalized placeholder', 'Y'),
('EBIT margin 2025', 0.040, 0.045, 0.032, '%', 'BHX scale + DMX recovery', 'Y'),
('EBIT margin 2026', 0.045, 0.052, 0.035, '%', '', 'Y'),
('EBIT margin 2027', 0.048, 0.056, 0.038, '%', '', 'Y'),
('EBIT margin 2028', 0.050, 0.058, 0.040, '%', '', 'Y'),
('EBIT margin 2029', 0.050, 0.060, 0.040, '%', '', 'Y'),
('Tax rate', 0.20, 0.20, 0.20, '%', 'VN corporate tax', 'Y'),
('D&A / Revenue', 0.015, 0.014, 0.016, '%', '', 'Y'),
('Capex / Revenue', 0.018, 0.016, 0.022, '%', 'Store expansion + IT/logistics', 'Y'),
('NWC / Revenue change', 0.08, 0.06, 0.10, '%', 'Working capital intensity', 'Y'),
('WACC', 0.105, 0.095, 0.115, '%', 'Cost of equity/debt placeholder', 'Y'),
('Terminal growth', 0.035, 0.040, 0.025, '%', 'Long-run VN nominal growth', 'Y'),
('Net debt', 6000, 4000, 8000, 'VNDbn', 'Debt - cash, placeholder', 'Y'),
('Shares outstanding', 1463, 1463, 1463, 'mn shares', 'Cập nhật theo báo cáo', 'Y'),
('TGDD EV/Sales', 0.35, 0.45, 0.25, 'x', 'Mobile retail mature', 'Y'),
('DMX EV/Sales pre-IPO', 0.45, 0.60, 0.35, 'x', 'Consumer electronics chain', 'Y'),
('BHX EV/Sales', 0.55, 0.80, 0.35, 'x', 'Grocery chain optionality', 'Y'),
('Other EV', 2000, 4000, 1000, 'VNDbn', 'An Khang/AVAKids/etc.', 'Y'),
('ĐMX IPO stake sold', 0.20, 0.25, 0.15, '%', 'Tỷ lệ bán trong IPO', 'Y'),
('ĐMX post-money valuation', 45000, 60000, 32000, 'VNDbn', 'Giả định định giá ĐMX khi IPO', 'Y'),
]
for r in assumptions: ws.append(r)
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.border=border
    row[1].fill=yellow; row[2].fill=green; row[3].fill=red

# Historical FS
ws = wb.create_sheet('Historical_FS')
cols = ['VNDbn','2020','2021','2022','2023','2024E/placeholder']
ws.append(cols)
for c in range(1,len(cols)+1): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
hist = [
('Revenue',108546,122958,133405,118280,134000),
('Gross profit',24600,27000,30100,24500,30000),
('EBIT',5600,6200,6400,1600,4700),
('NPAT-MI',3920,4900,4100,168,3600),
('Total assets',46000,59000,62000,60000,65000),
('Equity',17000,22000,25000,24500,27500),
('Net debt',5000,7000,6500,8000,6000),
('BHX revenue',21000,28000,30000,31500,40000),
('BHX EBIT margin',-0.08,-0.06,-0.04,-0.01,0.01),
]
for r in hist: ws.append(r)
for row in ws.iter_rows():
    for cell in row: cell.border=border
ws['A13']='Note'; ws['B13']='Các số lịch sử trong bản này là khung/placeholder để model chạy; cần cập nhật theo BCTC audited/quarterly chính thức của MWG.'

# DCF
ws = wb.create_sheet('DCF_Consolidated')
for col in range(1, 10): ws.column_dimensions[get_column_letter(col)].width=17
years = ['2024A','2025E','2026E','2027E','2028E','2029E','Terminal']
ws.append(['Line']+years)
for c in range(1,9): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
lines = ['Revenue','Growth','EBIT margin','EBIT','Tax','NOPAT','D&A','Capex','Change in NWC','FCFF','Discount factor','PV FCFF']
for i,l in enumerate(lines,2): ws.cell(i,1).value=l
# formulas
ws['B2']='=Assumptions!B2'
for col in range(3,8):
    ws.cell(2,col).value=f'={get_column_letter(col-1)}2*(1+Assumptions!B{col})'
ws['H2']='=G2*(1+Assumptions!B19)'
ws['B3']='';
for col in range(3,8): ws.cell(3,col).value=f'={col and get_column_letter(col)}2/{get_column_letter(col-1)}2-1'
for col in range(2,8): ws.cell(4,col).value=f'=Assumptions!B{6+col}' if col>=2 else ''
for col in range(2,8): ws.cell(5,col).value=f'={get_column_letter(col)}2*{get_column_letter(col)}4'
for col in range(2,8): ws.cell(6,col).value=f'={get_column_letter(col)}5*Assumptions!B14'
for col in range(2,8): ws.cell(7,col).value=f'={get_column_letter(col)}5-{get_column_letter(col)}6'
for col in range(2,8): ws.cell(8,col).value=f'={get_column_letter(col)}2*Assumptions!B15'
for col in range(2,8): ws.cell(9,col).value=f'=-{get_column_letter(col)}2*Assumptions!B16'
ws['B10']=0
for col in range(3,8): ws.cell(10,col).value=f'=-({get_column_letter(col)}2-{get_column_letter(col-1)}2)*Assumptions!B17'
for col in range(2,8): ws.cell(11,col).value=f'={get_column_letter(col)}7+{get_column_letter(col)}8+{get_column_letter(col)}9+{get_column_letter(col)}10'
for col in range(3,8): ws.cell(12,col).value=f'=1/(1+Assumptions!B18)^{col-2}'
for col in range(3,8): ws.cell(13,col).value=f'={get_column_letter(col)}11*{get_column_letter(col)}12'
ws['H11']='=G11*(1+Assumptions!B19)/(Assumptions!B18-Assumptions!B19)'
ws['H12']='=1/(1+Assumptions!B18)^5'
ws['H13']='=H11*H12'
ws['A15']='Enterprise value'; ws['B15']='=SUM(C13:H13)'
ws['A16']='Net debt'; ws['B16']='=Assumptions!B20'
ws['A17']='Equity value'; ws['B17']='=B15-B16'
ws['A18']='Shares mn'; ws['B18']='=Assumptions!B21'
ws['A19']='Value / share VND'; ws['B19']='=B17*1000/B18'
for row in ws.iter_rows():
    for cell in row: cell.border=border

# SOTP
ws = wb.create_sheet('SOTP')
ws.append(['Segment','Revenue assumption VNDbn','Multiple','EV VNDbn','Ownership','Equity attributable','Notes'])
for c in range(1,8): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
sotp = [
('TGDD/mobile',35000,'=Assumptions!B22','=B2*C2',1,'=D2*E2','Mature mobile retail, lower multiple'),
('ĐMX/consumer electronics',55000,'=Assumptions!B23','=B3*C3',1,'=D3*E3','IPO candidate; can crystallize value'),
('BHX/grocery',40000,'=Assumptions!B24','=B4*C4',1,'=D4*E4','Higher optionality if profitable scale'),
('Other/minority',0,'','=',1,'=Assumptions!B25','An Khang/AVAKids/etc.'),
]
for r in sotp: ws.append(r)
ws['A8']='Gross EV'; ws['F8']='=SUM(F2:F5)'
ws['A9']='Net debt'; ws['F9']='=Assumptions!B20'
ws['A10']='Equity value'; ws['F10']='=F8-F9'
ws['A11']='Value/share VND'; ws['F11']='=F10*1000/Assumptions!B21'
for row in ws.iter_rows():
    for cell in row: cell.border=border

# DMX IPO
ws = wb.create_sheet('DMX_IPO')
ws.append(['Item','Base formula/value','Interpretation'])
for c in range(1,4): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
items = [
('ĐMX post-money valuation VNDbn','=Assumptions!B27','Giá trị thị trường hàm ý của ĐMX sau IPO'),
('Stake sold','=Assumptions!B26','Tỷ lệ MWG bán ra/pha loãng'),
('Cash proceeds VNDbn','=B2*B3','Tiền về MWG/ĐMX tùy cấu trúc primary/secondary'),
('Remaining MWG ownership','=1-B3','Tỷ lệ MWG còn nắm ĐMX'),
('Value of retained ĐMX','=B2*B5','Giá trị phần còn lại MWG nắm giữ'),
('Potential debt reduction','=MIN(B4,Assumptions!B20)','Nếu dùng proceeds giảm nợ'),
('Equity value uplift vs pre-IPO SOTP','=B2-SOTP!D3','Chênh lệch giá trị ĐMX IPO so với multiple SOTP'),
]
for r in items: ws.append(r)
ws['A11']='Phân tích chuyên sâu'
ws['B11']='IPO ĐMX có thể tác động MWG qua 4 kênh: (1) crystallize value bằng market multiple riêng cho ĐMX; (2) tạo cash proceeds giúp giảm net debt/tài trợ BHX; (3) tăng minh bạch segment và mở khóa SOTP; (4) rủi ro dilution/control discount nếu bán tỷ lệ lớn hoặc market định giá thấp.'
ws['A13']='Tác động định giá'
ws['B13']='Trong DCF hợp nhất, IPO chủ yếu thay đổi capital structure và terminal optionality. Trong SOTP, IPO ĐMX tạo observable market value cho segment, thay vì gộp trong MWG multiple. Nếu post-money ĐMX > EV implied trong SOTP, fair value MWG tăng; ngược lại có thể tạo anchor thấp.'
for row in ws.iter_rows():
    for cell in row: cell.border=border

# BHX
ws = wb.create_sheet('BHX_Analysis')
ws.append(['Chủ đề','Nội dung phân tích','Implication định giá'])
for c in range(1,4): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
bhx = [
('Mô hình vận hành','Chuỗi grocery hiện đại: cửa hàng gần khu dân cư, SKU thiết yếu/tươi sống, logistics lạnh/khô, procurement tập trung, tối ưu shrinkage và vòng quay hàng tồn.','Khi đạt scale, gross margin + operating leverage cải thiện; valuation chuyển từ drag sang optionality.'),
('Revenue driver','Doanh thu/cửa hàng, số cửa hàng, mix hàng tươi sống/FMCG, online delivery, repeat frequency.','Sensitivity nên tách SSSG và store expansion.'),
('Cost driver','COGS/procurement, shrinkage, rent, labor/store, logistics, marketing, IT.','EBIT margin BHX là biến số quyết định upside.'),
('Thị phần','Thị trường grocery VN còn phân mảnh, kênh truyền thống lớn. BHX cạnh tranh với WinMart, Co.op, Satra, chợ truyền thống, minimart.','Nếu BHX lấy thị phần modern grocery và duy trì unit economics dương, multiple có thể cao hơn retail truyền thống.'),
('Rủi ro','Biên mỏng, hàng tươi sống hao hụt, cạnh tranh giá, logistics, mở rộng quá nhanh.','Bear case dùng EV/Sales thấp và margin hồi phục chậm.'),
('Triển vọng','Thu nhập đô thị tăng, nhu cầu tiện lợi/an toàn thực phẩm, thanh toán số, delivery hỗ trợ modern trade.','Bull case: doanh thu/cửa hàng tăng và margin ổn định.'),
]
for r in bhx: ws.append(r)
for row in ws.iter_rows():
    for cell in row: cell.border=border

# Macro
ws = wb.create_sheet('Macro_VN')
ws.append(['Yếu tố','Tác động đến MWG','Theo dõi'])
for c in range(1,4): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
macro = [
('GDP/thu nhập hộ gia đình','Tăng trưởng thu nhập hỗ trợ tiêu dùng ICT/CE và grocery modern trade.','GDP real, retail sales, CPI'),
('Lãi suất/tín dụng tiêu dùng','Lãi suất thấp hỗ trợ mua hàng điện máy/điện thoại trả góp.','Policy rate, consumer finance growth'),
('CPI thực phẩm','Ảnh hưởng sức mua và gross margin BHX.','Food CPI, commodity prices'),
('Tỷ giá','Ảnh hưởng hàng nhập khẩu electronics/phones.','USD/VND, vendor pricing'),
('Niềm tin tiêu dùng','Tác động chu kỳ thay mới điện thoại/điện máy.','Retail sales ex-inflation'),
]
for r in macro: ws.append(r)
for row in ws.iter_rows():
    for cell in row: cell.border=border

# Sensitivity
ws = wb.create_sheet('Sensitivity')
ws['A1']='DCF value/share sensitivity (VND)'; ws['A1'].font=bold
waccs=[0.09,0.10,0.105,0.11,0.12]; gs=[0.02,0.03,0.035,0.04,0.045]
ws.cell(2,1).value='WACC \ g'
for j,g in enumerate(gs,2): ws.cell(2,j).value=g
for i,w in enumerate(waccs,3):
    ws.cell(i,1).value=w
    for j,g in enumerate(gs,2):
        # approximate sensitivity references final year FCFF from DCF, recompute TV
        ws.cell(i,j).value=f'=((SUM(DCF_Consolidated!C13:G13)+(DCF_Consolidated!G11*(1+{g})/({w}-{g})/(1+{w})^5))-Assumptions!B20)*1000/Assumptions!B21'
for row in ws.iter_rows():
    for cell in row: cell.border=border

# Sources
ws = wb.create_sheet('Sources')
ws.append(['Nguồn','URL/ghi chú','Cần cập nhật'])
for c in range(1,4): ws.cell(1,c).fill=blue; ws.cell(1,c).font=white_font
sources=[
('MWG IR / annual reports','https://mwg.vn/ hoặc trang IR chính thức MWG','BCTC năm/quý, annual report, presentation'),
('HOSE/SSC disclosures','Công bố thông tin chính thức','Số cổ phiếu, nghị quyết IPO/spin-off nếu có'),
('CafeF/Vietstock','Market data, snapshot tài chính','Giá cổ phiếu, P/E, EV/EBITDA, historical FS cross-check'),
('General Statistics Office Vietnam','Macro GDP/CPI/retail sales','Macro assumptions'),
('Company store data','Số cửa hàng TGDD/ĐMX/BHX, doanh thu/cửa hàng','Segment assumptions'),
]
for r in sources: ws.append(r)
for row in ws.iter_rows():
    for cell in row: cell.border=border

# formatting all sheets
for ws in wb.worksheets:
    ws.freeze_panes='A2'
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = max(ws.column_dimensions[get_column_letter(col)].width or 12, 16)

wb.save(out)
print(out.resolve())
