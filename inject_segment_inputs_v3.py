from openpyxl import load_workbook
from pathlib import Path

p=Path('MWG_valuation_model_v3.xlsx')
wb=load_workbook(p)

# Historical consolidated placeholders refined toward realistic recent scale
ws=wb['Historical_5Y']
# rows mapping
hist = {
  2:[122958,133405,118280,134341,150000,'Revenue est/placeholder, 2025 raised above 2024 based on 2025 audited existence + recovery cycle'],
  3:[26757,28800,23600,28800,33000,'Gross profit placeholder refined'],
  4:[6600,6200,1600,4700,6200,'EBIT placeholder refined'],
  5:[4900,4100,168,3600,4700,'NPAT placeholder refined'],
  6:[8500,9000,6500,8000,9800,'Operating CF placeholder'],
  7:[2200,2500,1800,2200,2600,'Capex placeholder'],
  8:[14500,13000,12000,15000,17000,'Cash placeholder'],
  9:[21500,19500,20000,21000,24000,'Debt placeholder'],
 10:[7000,6500,8000,6000,7000,'Net debt placeholder'],
 11:[59000,62000,60000,65000,72000,'Assets placeholder'],
 12:[22000,25000,24500,27500,32000,'Equity placeholder'],
}
for r,vals in hist.items():
    for i,v in enumerate(vals[:5],2): ws.cell(r,i).value=v
    ws.cell(r,7).value=vals[5]

# Segment historical with manual inputs / assumptions grounded in business model
ws=wb['Segment_Historical']
seg = {
  2:[26000,26500,25000,25500,26000,'TGDD revenue - placeholder conservative, mature/saturated'],
  3:[900,900,880,850,820,'TGDD store count placeholder, optimization downtrend'],
  4:[28.9,29.4,28.4,30.0,31.7,'TGDD revenue/store = revenue/count approx'],
  5:[52000,56000,50000,56000,62000,'ĐMX revenue placeholder, major earnings engine'],
  6:[1600,1800,1850,1900,1950,'ĐMX store count placeholder'],
  7:[32.5,31.1,27.0,29.5,31.8,'ĐMX revenue/store approx'],
  8:[0.50,0.52,0.50,0.52,0.53,'ĐMX market share placeholder based on leading position'],
  9:[28000,30000,31500,40000,49000,'BHX revenue placeholder, fast growth'],
 10:[2000,2100,1700,1750,1950,'BHX store count placeholder after restructuring and reopening growth'],
 11:[14.0,14.3,18.5,22.9,25.1,'BHX revenue/store approx'],
 12:[-0.06,-0.04,-0.01,0.01,0.02,'BHX EBIT margin path placeholder: loss to near breakeven/profit'],
 13:[0.010,0.012,0.014,0.018,0.022,'BHX market share placeholder'],
 14:[44958,50905,42780,38841,39000,'Other revenue residual from consolidated less core segments'],
}
for r,vals in seg.items():
    for i,v in enumerate(vals[:5],2): ws.cell(r,i).value=v
    ws.cell(r,7).value=vals[5]

# Seed current / key assumptions more realistically
ws=wb['Assumptions']
updates = {
 2:70000,  # current price rough placeholder
 3:1463,
 4:7000,
 8:0.050,  # TGDD margin
 9:0.060,  # DMX margin
 10:0.040, # BHX target margin
 15:0.35,
 16:0.50,
 17:0.65,
 18:0.20,
 19:45000,
 20:0.20,
 21:0.05,
 22:0.10,
}
for row,val in updates.items(): ws.cell(row,2).value=val

# Current year base links in segment forecast sheets
# TGDD
ws=wb['TGDD_Forecast']
ws['B2']='=Segment_Historical!F3'
ws['B3']='=Segment_Historical!F4'
ws['B4']='=Segment_Historical!F2'
ws['B6']='=Assumptions!B8'
ws['B7']='=B4*B6'
# DMX
ws=wb['DMX_Forecast']
ws['B2']='=Segment_Historical!F6'
ws['B3']='=Segment_Historical!F8'
ws['B4']='=Segment_Historical!F7'
ws['B5']='=Segment_Historical!F5'
ws['B7']='=Assumptions!B9'
ws['B8']='=B5*B7'
# BHX
ws=wb['BHX_Forecast']
ws['B2']='=Segment_Historical!F10'
ws['B4']='=Segment_Historical!F11'
ws['B6']='=Segment_Historical!F9'
ws['B8']='=Segment_Historical!F12'
ws['B9']='=B6*B8'
ws['B10']='=Segment_Historical!F13'
# Others
ws=wb['Other_Subsidiaries']
ws['B2']=2500; ws['B3']=3500; ws['B4']=33000

# Add explicit note sheet for manual inputs rationale
if 'Manual_Input_Notes' in wb.sheetnames:
    del wb['Manual_Input_Notes']
ws=wb.create_sheet('Manual_Input_Notes')
ws.append(['Mục','Cách em input hiện tại','Mức độ tin cậy'])
notes=[
 ('TGDD/ĐMX/BHX revenue history','Ước lượng dựa trên quy mô tương đối của các mảng và tổng doanh thu MWG, dùng để model chạy chi tiết ngay.','Medium-low until official segment disclosures are keyed in'),
 ('BHX margin history','Đặt đường cong từ lỗ sang gần hòa vốn/lãi nhẹ để phản ánh turnaround.','Medium'),
 ('DMX market share','Assumption leader position ~50%+ in modern CE retail.','Medium-low'),
 ('Other revenue','Residual = consolidated less core segments.','Medium'),
 ('Mục tiêu','Cho anh có model chi tiết chạy được ngay; vòng sau nên thay bằng số từ annual report/presentation/BCTC note nếu trích được sạch.','High')
]
for r in notes: ws.append(r)
for col in ['A','B','C']: ws.column_dimensions[col].width=42

wb.save(p)
print(p.resolve())
