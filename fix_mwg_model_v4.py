from openpyxl import load_workbook
from pathlib import Path

src=Path('MWG_valuation_model_v4.xlsx')
out=Path('MWG_valuation_model_v4_fixed.xlsx')
wb=load_workbook(src)

# 1) Fix formulas that reference text/blank values, convert 2025 inputs into numbers only.
# Segment_Historical corrected 2025 values: keep as model inputs but mark estimate explicitly.
ws=wb['Segment_Historical']
# Conservative correction: 2025 total 150,000; segment split sums to total exactly.
# TGDD 25,000; DMX 65,000; BHX 45,000; Other 15,000 = 150,000
updates={
 2:[26000,26500,24000,25000,25000,'Estimate; replace with official MWG segment disclosure'],
 3:[900,880,850,820,800,'Estimate store optimization'],
 4:['=B2/B3','=C2/C3','=D2/D3','=E2/E3','=F2/F3','Formula'],
 5:[54000,60000,52000,60000,65000,'Estimate; ĐMX leading revenue contributor'],
 6:[1900,2000,2050,2100,2100,'Estimate'],
 7:['=B5/B6','=C5/C6','=D5/D6','=E5/E6','=F5/F6','Formula'],
 8:[0.50,0.51,0.50,0.52,0.53,'Estimate market share'],
 9:[28000,30000,31500,40000,45000,'Estimate; BHX growth but not overstate'],
 10:[2100,1800,1700,1750,1900,'Estimate'],
 11:['=B9/B10','=C9/C10','=D9/D10','=E9/E10','=F9/F10','Formula'],
 12:[-0.06,-0.04,-0.01,0.01,0.02,'Estimate EBIT margin path'],
 13:[0.010,0.012,0.014,0.018,0.021,'Estimate BHX market share'],
 14:[14958,16905,10780,9341,15000,'Residual = consolidated revenue - core segments'],
}
for r,vals in updates.items():
    for c,v in enumerate(vals[:5],2): ws.cell(r,c).value=v
    ws.cell(r,7).value=vals[5]

# Historical_5Y corrected consolidated 2025 to match segment total; label as estimate until audited tie-out
ws=wb['Historical_5Y']
updates={
 2:[122958,133405,118280,134341,150000,'Estimate placeholder; tie to CafeF/MWG audited 2025'],
 3:[26757,28800,23600,28800,33000,'Estimate placeholder'],
 4:[6600,6200,1600,4700,6200,'Estimate placeholder'],
 5:[4900,4100,168,3600,4700,'Estimate placeholder'],
 6:[8500,9000,6500,8000,9800,'Estimate placeholder'],
 7:[2200,2500,1800,2200,2600,'Estimate placeholder'],
 8:[14500,13000,12000,15000,17000,'Estimate placeholder'],
 9:[21500,19500,20000,21000,24000,'Estimate placeholder'],
 10:[7000,6500,8000,6000,7000,'Estimate placeholder'],
 11:[59000,62000,60000,65000,72000,'Estimate placeholder'],
 12:[22000,25000,24500,27500,32000,'Estimate placeholder'],
}
for r,vals in updates.items():
    for c,v in enumerate(vals[:5],2): ws.cell(r,c).value=v
    ws.cell(r,7).value=vals[5]

# Fix TGDD/DMX/BHX base formulas to correct row references
ws=wb['TGDD_Forecast']
ws['B2']='=Segment_Historical!F3'
ws['B3']='=Segment_Historical!F4'
ws['B4']='=Segment_Historical!F2'
ws['B6']='=Assumptions!B8'
ws['B7']='=B4*B6'
for col,prev in zip(['C','D','E','F','G'],['B','C','D','E','F']):
    ws[f'{col}2']=f'={prev}2*(1+Micro_Drivers!{col}3)'
    ws[f'{col}3']=f'={prev}3*(1+Macro_Drivers!{col}3+Micro_Drivers!{col}2)'
    ws[f'{col}4']=f'={col}2*{col}3'
    ws[f'{col}5']=f'=IFERROR({col}4/{prev}4-1,0)'
    ws[f'{col}6']='=Assumptions!B8'
    ws[f'{col}7']=f'={col}4*{col}6'

ws=wb['DMX_Forecast']
ws['B2']='=Segment_Historical!F6'
ws['B3']='=Segment_Historical!F8'
ws['B4']='=Segment_Historical!F7'
ws['B5']='=Segment_Historical!F5'
ws['B7']='=Assumptions!B9'
ws['B8']='=B5*B7'
for col,prev in zip(['C','D','E','F','G'],['B','C','D','E','F']):
    ws[f'{col}2']=f'={prev}2'
    ws[f'{col}3']=f'={prev}3+Micro_Drivers!{col}5'
    ws[f'{col}4']=f'={prev}4*(1+Macro_Drivers!{col}2+Macro_Drivers!{col}4+Micro_Drivers!{col}4)'
    ws[f'{col}5']=f'={col}2*{col}4*(1+({col}3-{prev}3))'
    ws[f'{col}6']=f'=IFERROR({col}5/{prev}5-1,0)'
    ws[f'{col}7']='=Assumptions!B9'
    ws[f'{col}8']=f'={col}5*{col}7'

ws=wb['BHX_Forecast']
ws['B2']='=Segment_Historical!F10'
ws['B3']=0
ws['B4']='=Segment_Historical!F11'
ws['B5']=0
ws['B6']='=Segment_Historical!F9'
ws['B8']='=Segment_Historical!F12'
ws['B9']='=B6*B8'
ws['B10']='=Segment_Historical!F13'
for col,prev in zip(['C','D','E','F','G'],['B','C','D','E','F']):
    ws[f'{col}2']=f'={prev}2+{col}3'
    ws[f'{col}3']=f'=Micro_Drivers!{col}6'
    ws[f'{col}4']=f'={prev}4*(1+Macro_Drivers!{col}5+Macro_Drivers!{col}6+Micro_Drivers!{col}7)'
    ws[f'{col}5']=f'=Micro_Drivers!{col}7'
    ws[f'{col}6']=f'={col}2*{col}4'
    ws[f'{col}7']=f'=IFERROR({col}6/{prev}6-1,0)'
    ws[f'{col}8']=f'=MIN(Assumptions!B10,{prev}8+0.007+Micro_Drivers!{col}8+Micro_Drivers!{col}9)'
    ws[f'{col}9']=f'={col}6*{col}8'
    ws[f'{col}10']=f'={prev}10+0.002'

# Fix Other_Subsidiaries base so no blanks create #VALUE
ws=wb['Other_Subsidiaries']
ws['B2']=2000; ws['B3']=3000; ws['B4']=10000
for row in range(2,5):
    ws[f'C{row}']=f'=B{row}*(1+Micro_Drivers!C10)'
    ws[f'D{row}']=f'=C{row}*(1+Micro_Drivers!D10)'
    ws[f'E{row}']=f'=D{row}*(1+Micro_Drivers!E10)'
    ws[f'F{row}']=f'=E{row}*(1+Micro_Drivers!F10)'
    ws[f'G{row}']=f'=F{row}*(1+Micro_Drivers!G10)'

# Fix Consolidated Forecast row 11 formula: SUMPRODUCT must not use text; use explicit EBIT from other with margin.
ws=wb['Consolidated_Forecast']
for col,prev in zip(['B','C','D','E','F','G'],['B','B','C','D','E','F']):
    ws[f'{col}2']=f'=TGDD_Forecast!{col}4'
    ws[f'{col}3']=f'=DMX_Forecast!{col}5'
    ws[f'{col}4']=f'=BHX_Forecast!{col}6'
    ws[f'{col}5']=f'=SUM(Other_Subsidiaries!{col}2:{col}4)'
    ws[f'{col}6']=f'=SUM({col}2:{col}5)'
    if col!='B': ws[f'{col}7']=f'=IFERROR({col}6/{prev}6-1,0)'
    ws[f'{col}8']=f'=TGDD_Forecast!{col}7'
    ws[f'{col}9']=f'=DMX_Forecast!{col}8'
    ws[f'{col}10']=f'=BHX_Forecast!{col}9'
    ws[f'{col}11']=f'=SUMPRODUCT(Other_Subsidiaries!{col}2:{col}4,Other_Subsidiaries!$I$2:$I$4)'
    ws[f'{col}12']=f'=SUM({col}8:{col}11)'
    ws[f'{col}13']=f'=IFERROR({col}12/{col}6,0)'
    ws[f'{col}14']=f'={col}6*Assumptions!B12'
    ws[f'{col}15']=f'=-{col}6*Assumptions!B13'
    ws[f'{col}16']=0 if col=='B' else f'=-({col}6-{prev}6)*Assumptions!B14'
    ws[f'{col}17']=f'={col}12*(1-Assumptions!B5)+{col}14+{col}15+{col}16'

# Fix DCF formulas: avoid array constants that older Excel/openpyxl may not calculate cleanly.
ws=wb['DCF']
ws['B2']='=SUM(B4:B8)'
ws['B3']='=Consolidated_Forecast!G17*(1+Assumptions!B6)/(Assumptions!B5-Assumptions!B6)'
ws['B4']='=Consolidated_Forecast!C17/(1+Assumptions!B5)^1'
ws['B5']='=Consolidated_Forecast!D17/(1+Assumptions!B5)^2'
ws['B6']='=Consolidated_Forecast!E17/(1+Assumptions!B5)^3'
ws['B7']='=Consolidated_Forecast!F17/(1+Assumptions!B5)^4'
ws['B8']='=Consolidated_Forecast!G17/(1+Assumptions!B5)^5'
ws['B9']='=B3/(1+Assumptions!B5)^5'
ws['B10']='=B2+B9'
ws['B11']='=B10-Assumptions!B4'
ws['B12']='=B11*1000/Assumptions!B3'

# Fix Dashboard links to DCF correct value/share row B12
ws=wb['Dashboard']
ws['B6']='=DCF!B12'
ws['B7']='=SOTP!F12'
ws['B8']='=IPO_DMX!B10'
ws['B9']='=B6*0.4+B7*0.4+B8*0.2'
ws['B10']='=IFERROR(B9/Assumptions!B2-1,0)'

# Fix Scenario output DCF row points
ws=wb['Scenario_Output']
ws['B2']='=DCF!B12*0.85'; ws['C2']='=DCF!B12'; ws['D2']='=DCF!B12*1.15'
ws['B3']='=SOTP!F12*0.85'; ws['C3']='=SOTP!F12'; ws['D3']='=SOTP!F12*1.20'
ws['B4']='=DMX_IPO_DeepDive!G2+SOTP!F12'; ws['C4']='=DMX_IPO_DeepDive!G3+SOTP!F12'; ws['D4']='=DMX_IPO_DeepDive!G4+SOTP!F12'
ws['B5']='=AVERAGE(B2:B4)'; ws['C5']='=AVERAGE(C2:C4)'; ws['D5']='=AVERAGE(D2:D4)'
ws['B6']='=B5/Assumptions!B2-1'; ws['C6']='=C5/Assumptions!B2-1'; ws['D6']='=D5/Assumptions!B2-1'

# Data quality update
ws=wb['Data_Quality']
ws.append(['2025 correction','Adjusted 2025 segment split to sum to consolidated estimate and marked as estimate, not official.','Medium-low','Need official MWG segment disclosure tie-out'])
ws.append(['Formula errors','Removed array formulas and blank-driven formulas causing #VALUE.','High','Open in Excel and recalc'])

wb.save(out)
print(out.resolve())
