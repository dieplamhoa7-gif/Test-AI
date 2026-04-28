from pathlib import Path
from openpyxl import load_workbook
import json
from datetime import datetime

src = Path('MWG_warrant_source.xlsm')
wb = load_workbook(src, data_only=True, keep_vba=True)
ws = wb.active
underlying = 'MWG'
current_underlying_price = ws['C1'].value
items = []
for col in [3,5,7,9,11]:  # C,E,G,I,K current columns
    code = ws.cell(3, col).value
    if not code:
        continue
    def val(row): return ws.cell(row, col).value
    maturity = val(7)
    if isinstance(maturity, datetime):
        maturity = maturity.strftime('%Y-%m-%d')
    price = val(4)
    conversion_ratio = val(5)
    exercise_price = val(6)
    days_left = val(8)
    leverage = val(10)
    breakeven = (exercise_price or 0) + (price or 0) * (conversion_ratio or 0)
    moneyness = 'ITM' if (current_underlying_price or 0) > (exercise_price or 0) else ('ATM' if abs((current_underlying_price or 0) - (exercise_price or 0)) / max(current_underlying_price or 1, 1) < 0.03 else 'OTM')
    items.append({
        'code': str(code),
        'underlying': underlying,
        'underlyingPrice': current_underlying_price,
        'fairValue': round(float(price or 0), 2),
        'conversionRatio': conversion_ratio,
        'exercisePrice': exercise_price,
        'maturityDate': maturity,
        'daysLeft': days_left,
        'leverage': round(float(leverage or 0), 2),
        'breakeven': round(float(breakeven or 0), 2),
        'moneyness': moneyness,
        'source': 'excel-upload'
    })
out = Path('app/warrants/warrants_static.json')
out.write_text(json.dumps({'updatedAt': datetime.now().isoformat(), 'items': items}, ensure_ascii=False, indent=2), encoding='utf-8')
print(out, len(items))
