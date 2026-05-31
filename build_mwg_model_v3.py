from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from pathlib import Path

out = Path('MWG_valuation_model_v3.xlsx')
wb = Workbook()

blue = PatternFill('solid', fgColor='1F4E78')
light = PatternFill('solid', fgColor='D9EAF7')
yellow = PatternFill('solid', fgColor='FFF2CC')
green = PatternFill('solid', fgColor='E2F0D9')
red = PatternFill('solid', fgColor='FCE4D6')
gray = PatternFill('solid', fgColor='E7E6E6')
white_font = Font(color='FFFFFF', bold=True)
bold = Font(bold=True)
small = Font(size=9)
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row=1, cols=None):
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = blue
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_all(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'B2'


def set_widths(ws, widths):
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

# README / dashboard
ws = wb.active
ws.title = 'Dashboard'
set_widths(ws, {'A':26,'B':18,'C':18,'D':18,'E':18,'F':18,'G':22})
ws['A1'] = 'MWG VALUATION MODEL V3 - DRIVER-BASED BY SEGMENT'
ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
ws['A3'] = 'Mục tiêu'
ws['B3'] = 'Forecast MWG theo từng ngành/mảng: TGDD (điện thoại), ĐMX (điện máy), BHX (grocery), mảng khác. Giả định liên kết với vĩ mô + vi mô doanh nghiệp.'
ws['A5'] = 'Kết quả định giá'
for r in range(6, 12):
    ws[f'A{r}'].fill = light
    ws[f'A{r}'].font = bold
    ws[f'A{r}'].border = border
    ws[f'B{r}'].border = border
ws['A6']='DCF value/share'; ws['B6']='=DCF!B10'
ws['A7']='SOTP value/share'; ws['B7']='=SOTP!E10'
ws['A8']='IPO ĐMX adjusted value/share'; ws['B8']='=IPO_DMX!B10'
ws['A9']='Blended target price'; ws['B9']='=B6*0.4+B7*0.4+B8*0.2'
ws['A10']='Upside/Downside vs current price'; ws['B10']='=IFERROR(B9/Assumptions!B32-1,"")'
ws['A12']='Trạng thái'
ws['B12']='Bản V3 đã tách forecast theo driver ngành + vi mô từng mảng. Cần cập nhật số lịch sử 2021-2025 và current price để ra output chính thức.'
style_all(ws)

# assumptions
ws = wb.create_sheet('Assumptions')
ws.append(['Input','Base','Bull','Bear','Unit','Ghi chú'])
style_header(ws,1,6)
assumptions = [
('Current MWG price', 68000, 68000, 68000, 'VND', 'Cập nhật giá thị trường hiện tại'),
('Shares outstanding', 1463, 1463, 1463, 'mn', 'Cập nhật theo báo cáo mới nhất'),
('Net debt 2025A', 7000, 5000, 9000, 'VNDbn', 'Debt - cash, cập nhật theo BCTC'),
('Tax rate', 0.20, 0.20, 0.20, '%', 'Thuế suất doanh nghiệp'),
('WACC', 0.105, 0.095, 0.115, '%', 'Chi phí vốn hợp nhất'),
('Terminal growth', 0.035, 0.040, 0.025, '%', 'Tăng trưởng dài hạn'),
('TGDD steady EBIT margin', 0.048, 0.053, 0.043, '%', 'Mảng mature, biên ổn định'),
('DMX steady EBIT margin', 0.060, 0.070, 0.050, '%', 'Điện máy lớn hơn, operating leverage tốt hơn'),
('BHX target EBIT margin', 0.040, 0.055, 0.020, '%', 'Khi BHX đạt scale bền vững'),
('Other segment EBIT margin', 0.015, 0.030, 0.000, '%', 'An Khang/AVAKids/others'),
('D&A / Revenue', 0.015, 0.014, 0.016, '%', 'Khấu hao tương đối ổn định'),
('Capex / Revenue', 0.018, 0.016, 0.022, '%', 'Capex duy trì + mở rộng'),
('NWC / Revenue delta', 0.080, 0.060, 0.100, '%', 'Nhu cầu vốn lưu động theo tăng trưởng'),
('TGDD EV/Sales', 0.35, 0.45, 0.25, 'x', 'Mảng retail mature'),
('DMX EV/Sales pre-IPO', 0.45, 0.60, 0.35, 'x', 'Điện máy before IPO rerating'),
('BHX EV/Sales', 0.55, 0.80, 0.35, 'x', 'Modern grocery optionality'),
('Other EV/Sales', 0.20, 0.30, 0.10, 'x', 'Other segments'),
('ĐMX IPO post-money valuation', 45000, 60000, 32000, 'VNDbn', 'Giả định định giá ĐMX khi IPO'),
('ĐMX IPO stake sold', 0.20, 0.25, 0.15, '%', 'Tỷ lệ IPO/spin-off'),
('BHX market share optionality premium', 0.05, 0.10, 0.00, 'x', 'Premium multiple khi BHX gain share mạnh'),
('Risk discount to holdco', 0.10, 0.05, 0.15, '%', 'Holding discount / execution risk')
]
for r in assumptions: ws.append(r)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):
    row[0].fill = yellow; row[1].fill = green; row[2].fill = red
set_widths(ws, {'A':34,'B':14,'C':14,'D':14,'E':12,'F':60})
style_all(ws)

# macro
ws = wb.create_sheet('Macro_Drivers')
forecast_years = ['2026E','2027E','2028E','2029E','2030E']
ws.append(['Macro factor'] + forecast_years + ['Why it matters'])
style_header(ws,1,7)
macro_rows = [
('VN nominal retail sales growth',0.10,0.09,0.08,0.07,0.07,'Anchor cho tổng cầu retail'),
('Disposable income growth',0.08,0.08,0.07,0.06,0.06,'Ảnh hưởng cầu điện thoại/điện máy'),
('Consumer confidence index proxy',0.04,0.04,0.03,0.03,0.02,'Ảnh hưởng goods ticket lớn'),
('Consumer finance / installment growth',0.09,0.08,0.07,0.06,0.06,'Hỗ trợ TGDD/ĐMX trả góp'),
('Food retail modern trade growth',0.14,0.13,0.12,0.11,0.10,'Anchor cho BHX'),
('Food CPI / pricing pass-through',0.04,0.04,0.035,0.03,0.03,'Tác động doanh thu danh nghĩa BHX'),
('USD/VND pressure',0.02,0.02,0.02,0.02,0.02,'Ảnh hưởng hàng nhập khẩu electronics')
]
for r in macro_rows: ws.append(r)
for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
    for c in row: c.fill = yellow
set_widths(ws, {'A':32,'B':12,'C':12,'D':12,'E':12,'F':12,'G':50})
style_all(ws)

# historical
ws = wb.create_sheet('Historical_5Y')
years = ['2021A','2022A','2023A','2024A','2025A']
ws.append(['VNDbn'] + years + ['Source/Note'])
style_header(ws,1,7)
for r in ['Revenue','Gross profit','EBIT','NPAT','Operating CF','Capex','Cash','Debt','Net debt','Total assets','Equity']:
    ws.append([r,'','','','','',''])
ws['A14']='Ghi chú'; ws['B14']='Điền số lịch sử 2021-2025 theo BCTC kiểm toán/hợp nhất MWG. Model forecast dùng 2025A làm base year.'
set_widths(ws, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':42})
style_all(ws)

# segment historical
ws = wb.create_sheet('Segment_Historical')
ws.append(['Metric'] + years + ['Source/Note'])
style_header(ws,1,7)
for r in [
    'TGDD revenue','TGDD store count','TGDD revenue/store',
    'DMX revenue','DMX store count','DMX revenue/store','DMX market share',
    'BHX revenue','BHX store count','BHX revenue/store','BHX EBIT margin','BHX market share',
    'Other revenue'
]:
    ws.append([r,'','','','','',''])
set_widths(ws, {'A':26,'B':14,'C':14,'D':14,'E':14,'F':14,'G':44})
style_all(ws)

# micro drivers by segment
ws = wb.create_sheet('Micro_Drivers')
ws.append(['Micro driver'] + forecast_years + ['Liên hệ với segment'])
style_header(ws,1,7)
rows = [
    ('TGDD same-store sales growth',0.04,0.04,0.03,0.03,0.02,'TGDD'),
    ('TGDD store optimization',0.00,0.00,0.00,0.00,0.00,'TGDD'),
    ('DMX same-store sales growth',0.06,0.05,0.04,0.04,0.03,'ĐMX'),
    ('DMX market share gain (ppt)',0.005,0.004,0.003,0.002,0.001,'ĐMX'),
    ('BHX new stores',250,220,200,180,150,'BHX'),
    ('BHX same-store sales growth',0.08,0.07,0.06,0.05,0.04,'BHX'),
    ('BHX shrinkage improvement (ppt)',0.003,0.002,0.002,0.001,0.001,'BHX'),
    ('BHX logistics efficiency uplift (ppt margin)',0.002,0.002,0.001,0.001,0.001,'BHX'),
    ('Other segment growth',0.10,0.08,0.06,0.05,0.05,'Others'),
]
for r in rows: ws.append(r)
for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
    for c in row: c.fill = yellow
set_widths(ws, {'A':36,'B':12,'C':12,'D':12,'E':12,'F':12,'G':24})
style_all(ws)

# TGDD forecast
ws = wb.create_sheet('TGDD_Forecast')
ws.append(['Line','2025A'] + forecast_years + ['Logic'])
style_header(ws,1,8)
for r in ['Store count','Revenue/store','Revenue','Growth','EBIT margin','EBIT']:
    ws.append([r])
# formulas base link
for col, prev in zip(['C','D','E','F','G'], ['B','C','D','E','F']):
    ws[f'{col}2'] = f'={prev}2*(1+Micro_Drivers!{col}3)'
    ws[f'{col}3'] = f'={prev}3*(1+Macro_Drivers!{col}3+Micro_Drivers!{col}2)'
    ws[f'{col}4'] = f'={col}2*{col}3'
    ws[f'{col}5'] = f'=IFERROR({col}4/{prev}4-1,"")'
    ws[f'{col}6'] = '=Assumptions!B8'
    ws[f'{col}7'] = f'={col}4*{col}6'
ws['B2']='=Segment_Historical!C2'; ws['B3']='=Segment_Historical!C3'; ws['B4']='=Segment_Historical!C4'; ws['B6']='=Assumptions!B8'; ws['B7']='=B4*B6'
for i, txt in enumerate(['TGDD phụ thuộc chu kỳ thay máy, thu nhập khả dụng, consumer finance và mức saturation của smartphone market.','Store count chủ yếu tối ưu, không mở rộng lớn như trước.'],2):
    ws[f'H{i}']=txt
set_widths(ws, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':54})
style_all(ws)

# DMX forecast
ws = wb.create_sheet('DMX_Forecast')
ws.append(['Line','2025A'] + forecast_years + ['Logic'])
style_header(ws,1,8)
for r in ['Store count','Market share','Revenue/store','Revenue','Growth','EBIT margin','EBIT']:
    ws.append([r])
for col, prev in zip(['C','D','E','F','G'], ['B','C','D','E','F']):
    ws[f'{col}2'] = f'={prev}2'
    ws[f'{col}3'] = f'={prev}3+Micro_Drivers!{col}5'
    ws[f'{col}4'] = f'={prev}4*(1+Macro_Drivers!{col}2+Macro_Drivers!{col}4+Micro_Drivers!{col}4)'
    ws[f'{col}5'] = f'={col}2*{col}4*(1+({col}3-{prev}3))'
    ws[f'{col}6'] = f'=IFERROR({col}5/{prev}5-1,"")'
    ws[f'{col}7'] = '=Assumptions!B9'
    ws[f'{col}8'] = f'={col}5*{col}7'
ws['B2']='=Segment_Historical!C5'; ws['B3']='=Segment_Historical!C8'; ws['B4']='=Segment_Historical!C7'; ws['B5']='=Segment_Historical!C6'; ws['B7']='=Assumptions!B9'; ws['B8']='=B5*B7'
ws['H2']='ĐMX link mạnh với hàng điện tử gia dụng, housing/household appliance cycle, thị phần và sức mua.'
set_widths(ws, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':54})
style_all(ws)

# BHX forecast
ws = wb.create_sheet('BHX_Forecast')
ws.append(['Line','2025A'] + forecast_years + ['Logic'])
style_header(ws,1,8)
for r in ['Store count','New stores','Revenue/store','Same-store sales growth','Revenue','Growth','EBIT margin','EBIT','Market share']:
    ws.append([r])
for col, prev in zip(['C','D','E','F','G'], ['B','C','D','E','F']):
    ws[f'{col}2'] = f'={prev}2+{col}3'
    ws[f'{col}3'] = f'=Micro_Drivers!{col}6'
    ws[f'{col}4'] = f'={prev}4*(1+Macro_Drivers!{col}5+Macro_Drivers!{col}6+Micro_Drivers!{col}7)'
    ws[f'{col}5'] = f'=Micro_Drivers!{col}7'
    ws[f'{col}6'] = f'={col}2*{col}4'
    ws[f'{col}7'] = f'=IFERROR({col}6/{prev}6-1,"")'
    ws[f'{col}8'] = f'=MIN(Assumptions!B10,{prev}8+0.007+Micro_Drivers!{col}8+Micro_Drivers!{col}9)'
    ws[f'{col}9'] = f'={col}6*{col}8'
    ws[f'{col}10'] = f'={prev}10+0.002'
ws['B2']='=Segment_Historical!C10'; ws['B3']=0; ws['B4']='=Segment_Historical!C11'; ws['B5']=0; ws['B6']='=Segment_Historical!C9'; ws['B8']='=Segment_Historical!C12'; ws['B9']='=B6*B8'; ws['B10']='=Segment_Historical!C13'
ws['H2']='BHX forecast theo số cửa hàng mở mới, doanh thu/cửa hàng, SSSG, food CPI, modern trade penetration và cải thiện logistics/shrinkage.'
set_widths(ws, {'A':24,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':58})
style_all(ws)

# other/subsidiaries
ws = wb.create_sheet('Other_Subsidiaries')
ws.append(['Business','2025A Revenue'] + forecast_years + ['2025A EBIT margin','Steady EBIT margin','Ghi chú'])
style_header(ws,1,10)
rows = [
    ('An Khang','', '', '', '', '', '', -0.03, 0.01, 'Nhà thuốc, scale/rationalization risk'),
    ('AVAKids / specialty retail','', '', '', '', '', '', -0.01, 0.02, 'Small-format specialty retail'),
    ('Other / online / services','', '', '', '', '', '', 0.03, 0.04, 'Gộp mảng nhỏ'),
]
for r in rows: ws.append(r)
for row in range(2,5):
    ws[f'C{row}']=f'=B{row}*(1+Micro_Drivers!C10)'
    ws[f'D{row}']=f'=C{row}*(1+Micro_Drivers!D10)'
    ws[f'E{row}']=f'=D{row}*(1+Micro_Drivers!E10)'
    ws[f'F{row}']=f'=E{row}*(1+Micro_Drivers!F10)'
    ws[f'G{row}']=f'=F{row}*(1+Micro_Drivers!G10)'
set_widths(ws, {'A':28,'B':16,'C':14,'D':14,'E':14,'F':14,'G':14,'H':16,'I':18,'J':42})
style_all(ws)

# consolidated forecast
ws = wb.create_sheet('Consolidated_Forecast')
ws.append(['Line','2025A'] + forecast_years + ['Logic'])
style_header(ws,1,8)
for r in ['TGDD revenue','DMX revenue','BHX revenue','Other revenue','Consolidated revenue','Growth','TGDD EBIT','DMX EBIT','BHX EBIT','Other EBIT','Consolidated EBIT','EBIT margin','D&A','Capex','Change NWC','FCFF']:
    ws.append([r])
for col, src in [('B','B'),('C','C'),('D','D'),('E','E'),('F','F'),('G','G')]:
    ws[f'{col}2']=f'=TGDD_Forecast!{src}4'
    ws[f'{col}3']=f'=DMX_Forecast!{src}5'
    ws[f'{col}4']=f'=BHX_Forecast!{src}6'
    ws[f'{col}5']=f'=SUM(Other_Subsidiaries!{src}2:{src}4)'
    ws[f'{col}6']=f'=SUM({col}2:{col}5)'
for prev, col in zip(['B','B','C','D','E','F'], ['B','C','D','E','F','G']):
    if col!='B': ws[f'{col}7']=f'=IFERROR({col}6/{prev}6-1,"")'
for col, src in [('B','B'),('C','C'),('D','D'),('E','E'),('F','F'),('G','G')]:
    ws[f'{col}8']=f'=TGDD_Forecast!{src}7'
    ws[f'{col}9']=f'=DMX_Forecast!{src}8'
    ws[f'{col}10']=f'=BHX_Forecast!{src}9'
    ws[f'{col}11']=f'=SUMPRODUCT(Other_Subsidiaries!${src}$2:${src}$4,Other_Subsidiaries!$I$2:$I$4)'
    ws[f'{col}12']=f'=SUM({col}8:{col}11)'
    ws[f'{col}13']=f'={col}12/{col}6'
    ws[f'{col}14']=f'={col}6*Assumptions!B12'
    ws[f'{col}15']=f'=-{col}6*Assumptions!B13'
    ws[f'{col}16']=0 if col=='B' else f'=-({col}6-{prev}6)*Assumptions!B14'
    ws[f'{col}17']=f'={col}12*(1-Assumptions!B5)+{col}14+{col}15+{col}16'
ws['H2']='Hợp nhất từ từng segment, không dùng 1 growth rate chung cho toàn MWG.'
set_widths(ws, {'A':24,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':58})
style_all(ws)

# DCF
ws = wb.create_sheet('DCF')
ws.append(['Item','Value','Ghi chú'])
style_header(ws,1,3)
rows = [
    ('PV FCFF 2026-2030','=SUM(C3:G3)',''),
    ('Terminal value','=Consolidated_Forecast!G17*(1+Assumptions!B6)/(Assumptions!B5-Assumptions!B6)','TV từ FCFF 2030'),
    ('PV FCFF by year','', 'helper row below'),
    ('2026 PV','=Consolidated_Forecast!C17/(1+Assumptions!B5)^1',''),
    ('2027 PV','=Consolidated_Forecast!D17/(1+Assumptions!B5)^2',''),
    ('2028 PV','=Consolidated_Forecast!E17/(1+Assumptions!B5)^3',''),
    ('2029 PV','=Consolidated_Forecast!F17/(1+Assumptions!B5)^4',''),
    ('2030 PV','=Consolidated_Forecast!G17/(1+Assumptions!B5)^5',''),
    ('PV Terminal','=B3/(1+Assumptions!B5)^5',''),
    ('Enterprise value','=SUM(B4:B8)+B9',''),
    ('Equity value','=B10-Assumptions!B4',''),
    ('Value/share','=B11*1000/Assumptions!B3','VND/share')
]
for r in rows: ws.append(r)
set_widths(ws, {'A':24,'B':22,'C':40})
style_all(ws)

# SOTP
ws = wb.create_sheet('SOTP')
ws.append(['Segment','2030 Revenue','Multiple','EV','Ownership / adj.','Equity value','Notes'])
style_header(ws,1,7)
rows = [
    ('TGDD','=TGDD_Forecast!G4','=Assumptions!B15','=B2*C2',1,'=D2*E2','Mature mobile retail'),
    ('ĐMX','=DMX_Forecast!G5','=Assumptions!B16','=B3*C3',1,'=D3*E3','Điện máy, optional IPO rerating'),
    ('BHX','=BHX_Forecast!G6','=Assumptions!B17+Assumptions!B21','=B4*C4',1,'=D4*E4','Modern grocery optionality'),
    ('Other','=SUM(Other_Subsidiaries!G2:G4)','=Assumptions!B18','=B5*C5',1,'=D5*E5','Mảng khác'),
]
for r in rows: ws.append(r)
ws['A8']='Gross EV'; ws['F8']='=SUM(F2:F5)'
ws['A9']='Net debt'; ws['F9']='=Assumptions!B4'
ws['A10']='Holdco discount'; ws['F10']='=F8*Assumptions!B22'
ws['A11']='Equity value'; ws['F11']='=F8-F9-F10'
ws['A12']='Value/share'; ws['F12']='=F11*1000/Assumptions!B3'
set_widths(ws, {'A':18,'B':16,'C':12,'D':16,'E':16,'F':16,'G':42})
style_all(ws)

# IPO DMX
ws = wb.create_sheet('IPO_DMX')
ws.append(['Item','Value','Interpretation'])
style_header(ws,1,3)
rows = [
    ('DMX IPO post-money','=Assumptions!B19','Giá trị thị trường giả định của ĐMX khi IPO'),
    ('Stake sold','=Assumptions!B20','Tỷ lệ bán ra'),
    ('Cash proceeds','=B2*B3','Tiền thu về'),
    ('Retained stake','=1-B3','Tỷ lệ còn nắm'),
    ('Retained value','=B2*B5','Giá trị phần còn nắm'),
    ('DMX base SOTP EV','=SOTP!D3','Giá trị ĐMX trước uplift IPO'),
    ('IPO uplift','=MAX(0,B2-B7)','Phần chênh lệch định giá nếu IPO > base'),
    ('Adj. equity value','=SOTP!F11+B8','Cộng uplift vào equity SOTP'),
    ('Value/share','=B9*1000/Assumptions!B3','VND/share'),
    ('Kết luận','IPO ĐMX nên được xem như công cụ mở khóa SOTP, giảm net debt/tài trợ BHX, và làm rõ market benchmark cho mảng điện máy.','Không nên chỉ nhìn như one-off gain.')
]
for r in rows: ws.append(r)
set_widths(ws, {'A':24,'B':18,'C':70})
style_all(ws)

# methodology
ws = wb.create_sheet('Methodology')
ws.append(['Mảng','Biến forecast chính','Macro link','Micro link','Định giá'])
style_header(ws,1,5)
rows = [
    ('TGDD','store count, revenue/store, SSSG, margin','thu nhập khả dụng, tín dụng tiêu dùng, confidence','chu kỳ thay máy, optimization cửa hàng','DCF + EV/Sales mature retail'),
    ('ĐMX','market share, revenue/store, margin','retail durables, housing/consumer cycle','share gain, assortment, financing','SOTP + IPO rerating logic'),
    ('BHX','new stores, revenue/store, SSSG, margin ramp','food retail growth, food CPI, urbanization','shrinkage, logistics, procurement, scale','EV/Sales + path to EBIT margin'),
    ('Others','doanh thu nhỏ, margin thấp','tùy thị trường ngách','execution discipline','conservative EV/Sales'),
]
for r in rows: ws.append(r)
set_widths(ws, {'A':16,'B':34,'C':28,'D':34,'E':26})
style_all(ws)

# sources
ws = wb.create_sheet('Sources')
ws.append(['Nguồn','Link / file','Ứng dụng'])
style_header(ws,1,3)
rows = [
    ('CafeF MWG BCTC page','tmp_mwg_cafef_financials.html','Nguồn BCTC và link PDF'),
    ('CafeF BCTC API list','tmp_mwg_bctc_list.json','Xác nhận đủ 2025 kiểm toán / 2026 Q1'),
    ('MWG 2025 audited consolidated PDF','MWG_BCTC_Hop_Nhat_2025_Kiem_Toan.pdf','Base year 2025A'),
    ('Macro assumptions','sheet Macro_Drivers','Gắn doanh thu forecast theo ngành'),
    ('Micro assumptions','sheet Micro_Drivers','Gắn riêng từng mảng')
]
for r in rows: ws.append(r)
set_widths(ws, {'A':28,'B':64,'C':38})
style_all(ws)

# charts
ws = wb['Dashboard']
chart = BarChart()
chart.title = 'Valuation Summary'
chart.y_axis.title = 'VND/share'
data = Reference(ws, min_col=2, max_col=2, min_row=6, max_row=8)
cats = Reference(ws, min_col=1, max_col=1, min_row=6, max_row=8)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
ws.add_chart(chart, 'D5')

for sh in wb.worksheets:
    for col in range(1, sh.max_column+1):
        letter = get_column_letter(col)
        if sh.column_dimensions[letter].width is None:
            sh.column_dimensions[letter].width = 16

wb.save(out)
print(out.resolve())
