from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

src=Path('MWG_valuation_model_v3.xlsx')
out=Path('MWG_valuation_model_v4.xlsx')
wb=load_workbook(src)
blue=PatternFill('solid', fgColor='1F4E78'); white=Font(color='FFFFFF',bold=True)
yellow=PatternFill('solid', fgColor='FFF2CC'); green=PatternFill('solid', fgColor='E2F0D9'); red=PatternFill('solid', fgColor='FCE4D6')
light=PatternFill('solid', fgColor='D9EAF7')
thin=Side(style='thin', color='D9D9D9'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def style(ws):
    for row in ws.iter_rows():
        for c in row:
            c.border=border; c.alignment=Alignment(vertical='top',wrap_text=True)
    ws.freeze_panes='B2'
def header(ws):
    for c in ws[1]: c.fill=blue; c.font=white; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
def widths(ws, d):
    for k,v in d.items(): ws.column_dimensions[k].width=v
for name in ['Segment_PnL','Peer_Multiples','Scenario_Output','BHX_Unit_Economics','DMX_IPO_DeepDive','Data_Quality']:
    if name in wb.sheetnames: del wb[name]
# Segment PnL detailed
ws=wb.create_sheet('Segment_PnL')
years=['2025A','2026E','2027E','2028E','2029E','2030E']
ws.append(['Line']+years+['Logic'])
header(ws)
rows=[
'TGDD Revenue','TGDD Gross margin','TGDD Gross profit','TGDD Opex/Sales','TGDD EBIT','TGDD EBIT margin',
'DMX Revenue','DMX Gross margin','DMX Gross profit','DMX Opex/Sales','DMX EBIT','DMX EBIT margin',
'BHX Revenue','BHX Gross margin','BHX Gross profit','BHX Shrinkage/Sales','BHX Logistics+Store Opex/Sales','BHX EBIT','BHX EBIT margin',
'Other Revenue','Other EBIT margin','Other EBIT','Total segment EBIT']
for r in rows: ws.append([r])
# year col mapping B:G to forecast sheets B:G
for col in ['B','C','D','E','F','G']:
    ws[f'{col}2']=f'=TGDD_Forecast!{col}4'; ws[f'{col}3']=0.205; ws[f'{col}4']=f'={col}2*{col}3'; ws[f'{col}5']=f'={col}3-TGDD_Forecast!{col}6'; ws[f'{col}6']=f'=TGDD_Forecast!{col}7'; ws[f'{col}7']=f'={col}6/{col}2'
    ws[f'{col}8']=f'=DMX_Forecast!{col}5'; ws[f'{col}9']=0.215; ws[f'{col}10']=f'={col}8*{col}9'; ws[f'{col}11']=f'={col}9-DMX_Forecast!{col}7'; ws[f'{col}12']=f'=DMX_Forecast!{col}8'; ws[f'{col}13']=f'={col}12/{col}8'
    ws[f'{col}14']=f'=BHX_Forecast!{col}6'; ws[f'{col}15']=0.245; ws[f'{col}16']=f'={col}14*{col}15';
    # shrinkage improves over time
    shrink={'B':0.025,'C':0.022,'D':0.020,'E':0.018,'F':0.017,'G':0.016}[col]
    ws[f'{col}17']=shrink; ws[f'{col}18']=f'={col}15-BHX_Forecast!{col}8-{col}17'; ws[f'{col}19']=f'=BHX_Forecast!{col}9'; ws[f'{col}20']=f'={col}19/{col}14'
    ws[f'{col}21']=f'=SUM(Other_Subsidiaries!{col}2:{col}4)'; ws[f'{col}22']=0.015; ws[f'{col}23']=f'={col}21*{col}22'; ws[f'{col}24']=f'=SUM({col}6,{col}12,{col}19,{col}23)'
ws['H2']='Tách gross margin / opex / shrinkage để thấy profit bridge từng mảng.'
widths(ws, {'A':34,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':60}); style(ws)
# Peer multiples
ws=wb.create_sheet('Peer_Multiples')
ws.append(['Peer/segment','Business type','EV/Sales','EV/EBITDA','P/E','Use for MWG','Note'])
header(ws)
peers=[
('FRT','ICT retail/pharmacy',0.45,12,22,'TGDD reference','VN listed peer, pharmacy mix differs'),
('DGW','ICT distribution',0.30,8,14,'TGDD/ICT sanity check','Distributor lower margin'),
('WinCommerce/Masan consumer retail','Grocery modern trade',0.70,18,30,'BHX reference','Grocery scale optionality'),
('Regional electronics retailer','Consumer electronics',0.50,10,18,'ĐMX reference','Cycle-sensitive'),
('MWG implied blended','Multi-format retail',0.45,10,18,'Cross-check','Depends BHX turnaround and IPO ĐMX')]
for r in peers: ws.append(r)
widths(ws, {'A':28,'B':30,'C':12,'D':12,'E':12,'F':22,'G':46}); style(ws)
# BHX unit economics
ws=wb.create_sheet('BHX_Unit_Economics')
ws.append(['Metric']+years+['Interpretation'])
header(ws)
metrics=['Store count','Revenue/store/year','Revenue/store/month','Gross margin','Shrinkage','Logistics+store opex','EBIT/store/year','EBIT margin','Market share']
for m in metrics: ws.append([m])
for col in ['B','C','D','E','F','G']:
    ws[f'{col}2']=f'=BHX_Forecast!{col}2'; ws[f'{col}3']=f'=BHX_Forecast!{col}4'; ws[f'{col}4']=f'={col}3/12'; ws[f'{col}5']=f'=Segment_PnL!{col}15'; ws[f'{col}6']=f'=Segment_PnL!{col}17'; ws[f'{col}7']=f'=Segment_PnL!{col}18'; ws[f'{col}8']=f'=BHX_Forecast!{col}9/{col}2'; ws[f'{col}9']=f'=BHX_Forecast!{col}8'; ws[f'{col}10']=f'=BHX_Forecast!{col}10'
ws['H2']='BHX phải xem theo unit economics: revenue/store, shrinkage, logistics/store opex, EBIT/store.'
widths(ws, {'A':28,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':60}); style(ws)
# IPO deep dive
ws=wb.create_sheet('DMX_IPO_DeepDive')
ws.append(['Case','IPO valuation','Stake sold','Proceeds','Retained value','Uplift vs base SOTP','Value/share impact','Comment'])
header(ws)
cases=[('Bear',32000,0.15),('Base',45000,0.20),('Bull',60000,0.25)]
for i,(case,val,stake) in enumerate(cases,2):
    ws.cell(i,1).value=case; ws.cell(i,2).value=val; ws.cell(i,3).value=stake; ws.cell(i,4).value=f'=B{i}*C{i}'; ws.cell(i,5).value=f'=B{i}*(1-C{i})'; ws.cell(i,6).value=f'=MAX(0,B{i}-SOTP!D3)'; ws.cell(i,7).value=f'=F{i}*1000/Assumptions!B3'; ws.cell(i,8).value='IPO creates observable valuation anchor; proceeds reduce balance-sheet pressure or fund BHX.'
widths(ws, {'A':14,'B':16,'C':12,'D':16,'E':16,'F':18,'G':18,'H':60}); style(ws)
# Scenarios
ws=wb.create_sheet('Scenario_Output')
ws.append(['Output','Bear','Base','Bull','Formula/logic'])
header(ws)
outs=[
('DCF value/share','=DCF!B12*0.85','=DCF!B12','=DCF!B12*1.15','Sensitivity wrapper'),
('SOTP value/share','=SOTP!F12*0.85','=SOTP!F12','=SOTP!F12*1.20','Multiple scenario'),
('IPO adjusted value/share','=DMX_IPO_DeepDive!G2+SOTP!F12','=DMX_IPO_DeepDive!G3+SOTP!F12','=DMX_IPO_DeepDive!G4+SOTP!F12','DMX IPO uplift'),
('Blended target price','=AVERAGE(B2:B4)','=AVERAGE(C2:C4)','=AVERAGE(D2:D4)','Equal-weight scenario'),
('Upside/downside','=B5/Assumptions!B2-1','=C5/Assumptions!B2-1','=D5/Assumptions!B2-1','vs current price')]
for r in outs: ws.append(r)
widths(ws, {'A':28,'B':18,'C':18,'D':18,'E':50}); style(ws)
# Data quality
ws=wb.create_sheet('Data_Quality')
ws.append(['Area','Current status','Confidence','Next action'])
header(ws)
rows=[
('Consolidated 2021-2025','Input manually to make model run; needs audit tie-out','Medium','Tie to BCTC statements'),
('TGDD revenue/store','Manual structured input','Medium-low','Pull from MWG annual report / monthly KPI'),
('DMX revenue/share','Manual structured input','Medium-low','Pull market share/company presentation'),
('BHX revenue/store/margin','Manual structured input with turnaround logic','Medium','Pull BHX disclosure, store count, monthly revenue/store'),
('IPO ĐMX','Scenario-based','Medium','Update when IPO prospectus/valuation appears')]
for r in rows: ws.append(r)
widths(ws, {'A':28,'B':48,'C':18,'D':58}); style(ws)
# update dashboard pointers
ws=wb['Dashboard']
ws['A14']='V4 additions'; ws['B14']='Segment_PnL, BHX_Unit_Economics, Peer_Multiples, DMX_IPO_DeepDive, Scenario_Output, Data_Quality.'
wb.save(out)
print(out.resolve())
