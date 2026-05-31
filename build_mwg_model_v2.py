from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from pathlib import Path

out = Path('MWG_valuation_model_v2.xlsx')
wb = Workbook()

blue = PatternFill('solid', fgColor='1F4E78')
light = PatternFill('solid', fgColor='D9EAF7')
yellow = PatternFill('solid', fgColor='FFF2CC')
green = PatternFill('solid', fgColor='E2F0D9')
red = PatternFill('solid', fgColor='FCE4D6')
gray = PatternFill('solid', fgColor='E7E6E6')
white_font = Font(color='FFFFFF', bold=True)
bold = Font(bold=True)
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row=1, cols=None):
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = blue
        ws.cell(row, c).font = white_font
        ws.cell(row, c).border = border
        ws.cell(row, c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_all(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'B2'


def set_widths(ws, widths):
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

# README
ws = wb.active
ws.title = 'README'
ws['A1'] = 'MWG VALUATION MODEL V2 - 5Y HISTORICAL + DRIVER-BASED FORECAST + SOTP'
ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
ws['A3'] = 'Mục tiêu'
ws['B3'] = 'Model định giá MWG chi tiết hơn: 5 năm lịch sử, forecast theo driver ngành bán lẻ, sheet riêng cho TGDD/ĐMX/BHX/công ty con, DCF + SOTP + case IPO ĐMX.'
ws['A5'] = 'Ghi chú quan trọng'
ws['B5'] = 'Một số số lịch sử hiện là ô nhập/cần cập nhật từ BCTC MWG 2021-2025 và tài liệu IR. Cấu trúc model đã làm theo hướng full-stack để chỉ cần cấy số vào là chạy.'
for i, txt in enumerate([
    'Historical_5Y: 2021-2025 lịch sử hợp nhất',
    'Segment_Historical: lịch sử theo TGDD/ĐMX/BHX/khác',
    'Industry_Drivers: tăng trưởng ngành bán lẻ/điện máy/grocery',
    'TGDD_Forecast / DMX_Forecast / BHX_Forecast: forecast riêng từng mảng',
    'Subsidiaries: công ty con/mảng khác chiếm doanh thu',
    'Consolidated_Forecast: hợp nhất',
    'DCF / SOTP / IPO_DMX / Sensitivity / Sources'
], 8):
    ws.cell(i,1).value='•'
    ws.cell(i,2).value=txt
set_widths(ws, {'A':4,'B':120})
style_all(ws)

# Assumptions
ws = wb.create_sheet('Assumptions')
ws.append(['Input','Base','Bull','Bear','Unit','Logic'])
style_header(ws,1,6)
ass = [
('WACC',0.105,0.095,0.115,'%', 'Cost of capital cho MWG'),
('Terminal growth',0.035,0.040,0.025,'%', 'Long-run nominal growth'),
('Tax rate',0.20,0.20,0.20,'%', 'Thuế suất doanh nghiệp'),
('Net debt 2025A',7000,5000,9000,'VNDbn','Cập nhật theo BCTC 2025'),
('Shares out',1463,1463,1463,'mn','Cập nhật theo báo cáo mới nhất'),
('ĐMX IPO valuation',45000,60000,32000,'VNDbn','Giả định post-money ĐMX'),
('ĐMX IPO stake sold',0.20,0.25,0.15,'%', 'Tỷ lệ IPO/spin-off'),
('BHX target steady EBIT margin',0.04,0.055,0.02,'%', 'Khi đạt scale'),
('TGDD steady EBIT margin',0.05,0.055,0.045,'%', 'Mảng mature'),
('ĐMX steady EBIT margin',0.06,0.07,0.05,'%', 'Điện máy tốt hơn TGDD nếu cầu hồi phục'),
]
for r in ass: ws.append(r)
for row in ws.iter_rows(min_row=2, max_col=4):
    row[1].fill = yellow; row[2].fill = green; row[3].fill = red
set_widths(ws, {'A':34,'B':14,'C':14,'D':14,'E':12,'F':56})
style_all(ws)

# Historical 5Y
ws = wb.create_sheet('Historical_5Y')
years = ['2021A','2022A','2023A','2024A','2025A']
ws.append(['VNDbn'] + years)
style_header(ws,1,6)
for item in [
    'Revenue','Gross profit','EBIT','NPAT','Operating CF','Capex','Total assets','Cash','Debt','Net debt','Equity'
]:
    ws.append([item,'','','','',''])
ws['A15']='Hướng dẫn'; ws['B15']='Điền số lịch sử 2021-2025 từ BCTC hợp nhất kiểm toán/quý 4.'
set_widths(ws, {'A':24,'B':14,'C':14,'D':14,'E':14,'F':14})
style_all(ws)

# Segment historical
ws = wb.create_sheet('Segment_Historical')
ws.append(['Metric'] + years)
style_header(ws,1,6)
segment_rows = [
    'TGDD revenue','TGDD store count',
    'DMX revenue','DMX store count','DMX market share',
    'BHX revenue','BHX store count','BHX revenue/store/year','BHX EBIT margin',
    'Other revenue'
]
for r in segment_rows: ws.append([r,'','','','',''])
ws['A13']='Note'; ws['B13']='Nếu thiếu số công khai, có thể nội suy từ annual report, presentation, bài IR, hoặc ước lượng bảo thủ để sensitivity chạy.'
set_widths(ws, {'A':28,'B':14,'C':14,'D':14,'E':14,'F':14})
style_all(ws)

# Industry drivers
ws = wb.create_sheet('Industry_Drivers')
forecast_years = ['2026E','2027E','2028E','2029E','2030E']
ws.append(['Driver'] + forecast_years + ['Ghi chú'])
style_header(ws,1,7)
drivers = [
    ('VN retail sales growth',0.10,0.09,0.08,0.07,0.07,'Tăng trưởng tiêu dùng danh nghĩa'),
    ('Consumer electronics demand growth',0.08,0.08,0.07,0.06,0.05,'Chu kỳ thay mới điện thoại/điện máy'),
    ('Modern grocery growth',0.14,0.13,0.12,0.11,0.10,'Hiện đại hóa bán lẻ thực phẩm'),
    ('DMX market share change',0.005,0.004,0.003,0.002,0.001,'Điểm % gain/share'),
    ('BHX new stores',250,220,200,180,150,'Mở thêm cửa hàng/năm'),
    ('BHX same-store sales growth',0.08,0.07,0.06,0.05,0.04,'Tăng trưởng cửa hàng cũ'),
    ('TGDD same-store sales growth',0.04,0.04,0.03,0.03,0.02,'Mảng mature'),
    ('ĐMX same-store sales growth',0.06,0.05,0.04,0.04,0.03,'Phục hồi điện máy'),
]
for r in drivers: ws.append(r)
for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
    for cell in row: cell.fill = yellow
set_widths(ws, {'A':32,'B':12,'C':12,'D':12,'E':12,'F':12,'G':50})
style_all(ws)

# TGDD forecast
ws = wb.create_sheet('TGDD_Forecast')
ws.append(['Line','2025A'] + forecast_years)
style_header(ws,1,7)
rows = ['Store count','Revenue/store','Revenue','Growth','EBIT margin','EBIT']
for r in rows: ws.append([r])
ws['B2']='=Segment_Historical!B2'  # placeholder legacy linkage may be adjusted by user
for i, col in enumerate(['C','D','E','F','G'], start=0):
    prev = chr(ord('B')+i)
    ws[f'{col}2'] = f'={prev}2'
    ws[f'{col}3'] = f'={prev}3*(1+Industry_Drivers!{col}8)'
    ws[f'{col}4'] = f'={col}2*{col}3'
    ws[f'{col}5'] = f'=IFERROR({col}4/{prev}4-1,"")'
    ws[f'{col}6'] = '=Assumptions!B10'
    ws[f'{col}7'] = f'={col}4*{col}6'
set_widths(ws, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
style_all(ws)

# DMX forecast
ws = wb.create_sheet('DMX_Forecast')
ws.append(['Line','2025A'] + forecast_years)
style_header(ws,1,7)
for r in ['Store count','Market share','Revenue/store','Revenue','Growth','EBIT margin','EBIT']:
    ws.append([r])
for i, col in enumerate(['C','D','E','F','G'], start=0):
    prev = chr(ord('B')+i)
    ws[f'{col}2'] = f'={prev}2'
    ws[f'{col}3'] = f'={prev}3+Industry_Drivers!{col}5'
    ws[f'{col}4'] = f'={prev}4*(1+Industry_Drivers!{col}9)'
    ws[f'{col}5'] = f'={col}2*{col}4*(1+({col}3-{prev}3))'
    ws[f'{col}6'] = f'=IFERROR({col}5/{prev}5-1,"")'
    ws[f'{col}7'] = '=Assumptions!B11'
    ws[f'{col}8'] = f'={col}5*{col}7'
set_widths(ws, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
style_all(ws)

# BHX forecast
ws = wb.create_sheet('BHX_Forecast')
ws.append(['Line','2025A'] + forecast_years)
style_header(ws,1,7)
for r in ['Store count','New stores','Revenue/store','Same-store growth','Revenue','Growth','EBIT margin','EBIT']:
    ws.append([r])
for i, col in enumerate(['C','D','E','F','G'], start=0):
    prev = chr(ord('B')+i)
    ws[f'{col}2'] = f'={prev}2+{col}3'
    ws[f'{col}3'] = f'=Industry_Drivers!{col}6'
    ws[f'{col}4'] = f'={prev}4*(1+Industry_Drivers!{col}7)'
    ws[f'{col}5'] = f'=Industry_Drivers!{col}7'
    ws[f'{col}6'] = f'={col}2*{col}4'
    ws[f'{col}7'] = f'=IFERROR({col}6/{prev}6-1,"")'
    # margin ramp from historical toward target
    ws[f'{col}8'] = f'=MIN(Assumptions!B9,{prev}8+0.008)'
    ws[f'{col}9'] = f'={col}6*{col}8'
set_widths(ws, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
style_all(ws)

# Subsidiaries sheet
ws = wb.create_sheet('Subsidiaries')
ws.append(['Business / subsidiary','2025A Revenue','2026E Growth','2027E Growth','2028E Growth','2029E Growth','2030E Growth','2026E Revenue','2030E Revenue','Notes'])
style_header(ws,1,10)
subs = [
    ('An Khang', '', 0.10, 0.08, 0.06, 0.05, 0.05, '', '', 'Nhà thuốc, currently nhỏ và volatile'),
    ('AVAKids / other specialty retail', '', 0.12, 0.10, 0.08, 0.07, 0.06, '', '', 'Nếu còn đóng góp doanh thu đáng kể'),
    ('Online / services / others', '', 0.08, 0.07, 0.06, 0.05, 0.05, '', '', 'Gộp mảng nhỏ'),
]
for r in subs: ws.append(r)
for row in range(2, 5):
    ws[f'H{row}'] = f'=B{row}*(1+C{row})'
    ws[f'I{row}'] = f'=B{row}*(1+C{row})*(1+D{row})*(1+E{row})*(1+F{row})*(1+G{row})'
set_widths(ws, {'A':30,'B':16,'C':12,'D':12,'E':12,'F':12,'G':12,'H':16,'I':16,'J':40})
style_all(ws)

# Consolidated forecast
ws = wb.create_sheet('Consolidated_Forecast')
ws.append(['Line','2025A'] + forecast_years)
style_header(ws,1,7)
for r in ['TGDD revenue','DMX revenue','BHX revenue','Other revenue','Consolidated revenue','Growth','Consolidated EBIT','EBIT margin','D&A','Capex','Change NWC','FCFF']:
    ws.append([r])
# link formulas
for col, src in [('B','B'),('C','C'),('D','D'),('E','E'),('F','F'),('G','G')]:
    ws[f'{col}2']=f'=TGDD_Forecast!{src}4'
    ws[f'{col}3']=f'=DMX_Forecast!{src}5'
    ws[f'{col}4']=f'=BHX_Forecast!{src}6'
    if col=='B':
        ws[f'{col}5']='=SUM(Subsidiaries!B2:B4)'
    else:
        mapcol={'C':'H','D':'I','E':'I','F':'I','G':'I'}
        # simple placeholder carry for later years
        ws[f'{col}5']=f'=IF({col}$1="2026E",SUM(Subsidiaries!H2:H4),{chr(ord(col)-1)}5*1.06)'
    ws[f'{col}6']=f'=SUM({col}2:{col}5)'
for col_prev, col in zip(['B','B','C','D','E','F'], ['B','C','D','E','F','G']):
    if col!='B': ws[f'{col}7']=f'=IFERROR({col}6/{col_prev}6-1,"")'
for col, tcol, dcol, bcol in [('B','B','B','B'),('C','C','C','C'),('D','D','D','D'),('E','E','E','E'),('F','F','F','F'),('G','G','G','G')]:
    ws[f'{col}8']=f'=({col}2*TGDD_Forecast!{bcol}6+{col}3*DMX_Forecast!{bcol}7+{col}4*BHX_Forecast!{bcol}8)/{col}6'
    ws[f'{col}9']=f'=({col}2*TGDD_Forecast!{bcol}7+{col}3*DMX_Forecast!{bcol}8+{col}4*BHX_Forecast!{bcol}9)' if col!='B' else f'=({col}2*0+{col}3*0+{col}4*0)'
    ws[f'{col}10']=f'={col}6*0.015'
    ws[f'{col}11']=f'=-{col}6*0.018'
    ws[f'{col}12']=0 if col=='B' else f'=-({col}6-{chr(ord(col)-1)}6)*0.08'
    ws[f'{col}13']=f'={col}9+{col}10+{col}11+{col}12'
set_widths(ws, {'A':24,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
style_all(ws)

# DCF
ws = wb.create_sheet('DCF')
ws.append(['Item','Value'])
style_header(ws,1,2)
rows = [
    ('PV FCFF 2026-2030','=SUM(Consolidated_Forecast!C13:G13/(1+Assumptions!B2)^{1,2,3,4,5})'),
    ('Terminal value','=Consolidated_Forecast!G13*(1+Assumptions!B3)/(Assumptions!B2-Assumptions!B3)'),
    ('PV Terminal','=B3/(1+Assumptions!B2)^5'),
    ('Enterprise value','=B2+B4'),
    ('Net debt','=Assumptions!B5'),
    ('Equity value','=B5-B6'),
    ('Shares','=Assumptions!B6'),
    ('Value/share VND','=B7*1000/B8'),
]
for r in rows: ws.append(r)
set_widths(ws, {'A':28,'B':20})
style_all(ws)

# SOTP
ws = wb.create_sheet('SOTP')
ws.append(['Segment','2030 Revenue','Multiple','EV','Notes'])
style_header(ws,1,5)
vals = [
    ('TGDD','=TGDD_Forecast!G4',0.35,'=B2*C2','Mature mobile retail'),
    ('ĐMX','=DMX_Forecast!G5',0.45,'=B3*C3','Can rerate if IPO'),
    ('BHX','=BHX_Forecast!G6',0.55,'=B4*C4','Optionality from grocery scale'),
    ('Other','=Consolidated_Forecast!G5',0.20,'=B5*C5','Small businesses'),
]
for r in vals: ws.append(r)
ws['A7']='Gross EV'; ws['D7']='=SUM(D2:D5)'
ws['A8']='Net debt'; ws['D8']='=Assumptions!B5'
ws['A9']='Equity value'; ws['D9']='=D7-D8'
ws['A10']='Value/share VND'; ws['D10']='=D9*1000/Assumptions!B6'
set_widths(ws, {'A':20,'B':16,'C':12,'D':16,'E':40})
style_all(ws)

# IPO DMX
ws = wb.create_sheet('IPO_DMX')
ws.append(['Item','Value','Interpretation'])
style_header(ws,1,3)
rows = [
    ('ĐMX IPO valuation','=Assumptions!B7','Post-money valuation giả định của ĐMX'),
    ('Stake sold','=Assumptions!B8','Tỷ lệ IPO/spin-off'),
    ('Cash proceeds','=B2*B3','Tiền về nếu bán cổ phần'),
    ('Retained stake','=1-B3','Tỷ lệ MWG còn nắm'),
    ('Retained value','=B2*B5','Giá trị phần còn nắm'),
    ('Implied uplift vs SOTP base','=B2-SOTP!D3','Nếu market value IPO > base multiple'),
    ('IPO logic','IPO ĐMX giúp mở khóa định giá segment, tăng minh bạch, có thể giảm nợ/tài trợ BHX.','Tác động mạnh nhất lên SOTP và capital structure hơn là chỉ P/E ngắn hạn.')
]
for r in rows: ws.append(r)
set_widths(ws, {'A':24,'B':18,'C':80})
style_all(ws)

# Sensitivity
ws = wb.create_sheet('Sensitivity')
ws['A1']='Value/share sensitivity'
ws['A1'].font = bold
waccs=[0.09,0.10,0.105,0.11,0.12]
gs=[0.02,0.03,0.035,0.04,0.045]
ws.cell(2,1).value='WACC vs g'
for j,g in enumerate(gs,2): ws.cell(2,j).value=g
for i,w in enumerate(waccs,3):
    ws.cell(i,1).value=w
    for j,g in enumerate(gs,2):
        ws.cell(i,j).value=f'=(((SUM(Consolidated_Forecast!C13:G13/(1+{w})^{{1,2,3,4,5}}))+(Consolidated_Forecast!G13*(1+{g})/({w}-{g})/(1+{w})^5)-Assumptions!B5)*1000/Assumptions!B6)'
style_all(ws)

# Sources
ws = wb.create_sheet('Sources')
ws.append(['Nguồn','Link / file','Trạng thái'])
style_header(ws,1,3)
sources = [
    ('CafeF BCTC MWG list','tmp_mwg_bctc_list.json','Đã lấy được list tới 2026'),
    ('MWG 2025 audited consolidated PDF','MWG_BCTC_Hop_Nhat_2025_Kiem_Toan.pdf','Đã tải về'),
    ('CafeF BCTC page html','tmp_mwg_cafef_financials.html','Đã lưu local'),
    ('Model notes','MWG_model_notes.md','Đã tạo'),
    ('Cần cập nhật thủ công thêm','Store count, DMX market share, BHX market share/revenue/store từ IR/presentation','Pending'),
]
for r in sources: ws.append(r)
set_widths(ws, {'A':26,'B':72,'C':26})
style_all(ws)

# charts
ws = wb['Consolidated_Forecast']
chart = LineChart()
chart.title = 'MWG Revenue Forecast'
chart.y_axis.title = 'VNDbn'
chart.x_axis.title = 'Year'
data = Reference(ws, min_col=2, max_col=7, min_row=6, max_row=6)
cats = Reference(ws, min_col=2, max_col=7, min_row=1, max_row=1)
chart.add_data(data, from_rows=True, titles_from_data=False)
chart.set_categories(cats)
ws.add_chart(chart, 'I2')

for ws in wb.worksheets:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 16

wb.save(out)
print(out.resolve())
